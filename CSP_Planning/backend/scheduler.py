"""
Description : Planificateur de répétitions pour l'Orchestraaaaa
Licence : On devrait peut-être mettre une licence hein
Anno : 43
Auteur : Mateo Bauvir
"""
import pandas as pd
import random
import time
from collections import defaultdict, Counter
from typing import Dict, List, Set, Tuple, Optional
import re

class OptimizedRepetitionScheduler:
    def __init__(self,
                 repartitions_file: str,
                 disponibilites_file: str,
                 maybe_penalty: int = 10,
                 max_load: int = 3,
                 load_penalty: int = 50,
                 group_bonus: int = 20,
                 mode_absence: str = "flexible",
                 seuil_absence: int = 2,
                 generation_time_limit: int = 30):
        
        self.repartitions_file = repartitions_file
        self.disponibilites_file = disponibilites_file
        self.maybe_penalty = maybe_penalty
        self.max_load = max_load
        self.load_penalty = load_penalty
        self.group_bonus = group_bonus
        self.mode_absence = mode_absence
        self.seuil_absence = seuil_absence
        self.generation_time_limit = generation_time_limit
        
        self.musiciens: Set[str] = set()
        self.morceaux: List[str] = []
        self.creneaux: List[str] = []
        self.repartition: Dict[str, Set[str]] = {}  # morceau -> musiciens
        self.repartitions_df = None
        self.disponibilites_df = None
        self.disponibilites: Dict[str, Dict[str, str]] = {}  # musicien -> {creneau: dispo}
        self.creneaux_par_jour: Dict[str, List[str]] = defaultdict(list)
        self.slot_index: Dict[str, int] = {}
        self.musiciens_absents_force: Dict[str, Set[str]] = defaultdict(set)
        self.absent_participants: Dict[str, Dict[str, List]] = defaultdict(lambda: defaultdict(list))
        
        # Variables CSP Min-Conflicts
        self.assignment: Dict[str, Optional[str]] = {}  # morceau -> creneau (ou None)
        self.conflicts: Dict[str, int] = {}  # morceau -> nombre de conflits
        self.solution: Dict[str, str] = {}          # Solution finale (morceau -> creneau)
        self.status = None  
        
        # Cache pour optimisation
        self._conflict_cache: Dict[Tuple[str, str], int] = {}
        self._musicien_morceaux: Dict[str, List[str]] = defaultdict(list)
        
        # Paramètres algorithme
        self.max_iterations = 10000
        self.max_restarts = generation_time_limit
        
    def transformer_simple(self, texte: str) -> Optional[Dict]:
        """Transforme un texte de créneau en dictionnaire structuré."""
        t = texte.strip().replace("\n", " ").replace("\r", " ")
        m = re.search(r"(\w+\.)\s+(\d+).*?(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})", t)
        if not m:
            return None
        
        jour_txt, jour_chiffre, H1, M1, H2, M2 = m.groups()
        jours = {
            'lun.': 'LUN', 'mar.': 'MAR', 'mer.': 'MER',
            'jeu.': 'JEU', 'ven.': 'VEN', 'sam.': 'SAM', 'dim.': 'DIM'
        }
        jour = jours.get(jour_txt.lower(), jour_txt.upper().rstrip('.'))
        
        return {
            'jour': jour,
            'date': int(jour_chiffre),
            'h1': int(H1), 'm1': int(M1),
            'h2': int(H2), 'm2': int(M2)
        }
    
    def load_data(self):
        """Charge les données depuis les fichiers Excel."""
        # print("Chargement des données...")
        self.repartitions_df = pd.read_excel(self.repartitions_file)
        self.disponibilites_df = pd.read_excel(self.disponibilites_file)
        instrument_cols = self.repartitions_df.columns[6:]
        
        for _, row in self.repartitions_df.iterrows():
            morceau = row['Titre']
            if pd.isna(morceau):
                continue
                
            has_musicians = any(not pd.isna(row[c]) for c in instrument_cols)
            if not has_musicians:
                continue
                
            self.morceaux.append(morceau)
            self.repartition[morceau] = set()
            
            for col in instrument_cols:
                cellule = row[col]
                if pd.isna(cellule):
                    continue
                    
                for nom in str(cellule).split(','):
                    nom = nom.strip()
                    if nom:
                        self.musiciens.add(nom)
                        self.repartition[morceau].add(nom)
                        self._musicien_morceaux[nom].append(morceau)
        
        dispo_cols = self.disponibilites_df.columns[2:]
        all_dates = set()
        
        for _, row in self.disponibilites_df.iterrows():
            musicien = str(row['Nom']).strip().title()
            if pd.isna(musicien) or not musicien:
                continue
                
            self.disponibilites[musicien] = {}
            
            for col in dispo_cols:
                info = self.transformer_simple(str(col))
                if not info:
                    continue
                    
                d = info['date']
                j = info['jour']
                h1, m1, h2, m2 = info['h1'], info['m1'], info['h2'], info['m2']
                all_dates.add(d)
                start = f"{h1:02d}:{m1:02d}"
                end = f"{h2:02d}:{m2:02d}"
                slot = f"{j}_{d:02d}_{start}-{end}"
                
                val = str(row[col]).strip().lower() if not pd.isna(row[col]) else "non"
                self.disponibilites[musicien][slot] = val
        
        if self.disponibilites:
            premier = next(iter(self.disponibilites.values()))
            
            def _key(s):
                _, dd, plage = s.split("_")
                start, _ = plage.split("-")
                h, m = start.split(":")
                return (int(dd), int(h), int(m))
            
            self.creneaux = sorted(premier.keys(), key=_key)
            self.slot_index = {slot: i for i, slot in enumerate(self.creneaux)}
            
            # Regroupement par jour
            for slot in self.creneaux:
                jour = slot.split("_")[0]
                self.creneaux_par_jour[jour].append(slot)
        
        dates_sorted = sorted(all_dates)
        if dates_sorted:
            base = dates_sorted[0]
            self.date2week = {d: ((d - base) // 7) + 1 for d in dates_sorted}
            self.weeks = sorted(set(self.date2week.values()))
        
        print(f"Données chargées: {len(self.morceaux)} morceaux, {len(self.musiciens)} musiciens, {len(self.creneaux)} créneaux")
    
    def build_model(self):
        """Construit le modèle CSP (remplace la construction OR-Tools)."""
        print("Construction du modèle CSP optimisé...")
        
        # Réinitialisation
        self.assignment = {}
        self.conflicts = {}
        self._conflict_cache.clear()
        self.musiciens_absents_force.clear()
        self.absent_participants.clear()
        for morceau in self.morceaux:
            self.assignment[morceau] = None
            self.conflicts[morceau] = 0
    
    def calculate_conflicts(self, morceau: str, creneau: str) -> int:
        """Calcule le nombre de conflits pour assigner un morceau à un créneau."""
        # Vérifier le cache
        cache_key = (morceau, creneau)
        if cache_key in self._conflict_cache:
            return self._conflict_cache[cache_key]
        
        conflicts = 0
        musiciens_morceau = self.repartition[morceau]
        
        # 1. Conflits de disponibilité
        absents = 0
        maybe_count = 0
        
        for musicien in musiciens_morceau:
            dispo = self.disponibilites.get(musicien, {}).get(creneau, "non")
            if dispo == "non":
                absents += 1
                if self.mode_absence == "strict":
                    conflicts += 10000  # Pénalité forte en mode strict
                else:
                    conflicts += 100   # Pénalité modérée en mode flexible
            elif dispo == "peut-être":
                maybe_count += 1
                conflicts += self.maybe_penalty
        
        # Gestion du seuil d'absences
        if self.mode_absence != "strict" and absents > self.seuil_absence:
            conflicts += (absents - self.seuil_absence) * 1000
        
        # 2. Conflit de créneau (un seul morceau par créneau)
        for autre_morceau, autre_creneau in self.assignment.items():
            if autre_morceau != morceau and autre_creneau == creneau:
                conflicts += 100000000  # sinon on aurait des double assignation
        
        # 3. Conflits de charge quotidienne
        jour = creneau.split("_")[0]
        for musicien in musiciens_morceau:
            charge_jour = self._get_daily_load(musicien, jour, exclude_morceau=morceau)
            
            # Simuler l'ajout du créneau
            if any(self.assignment.get(m) == creneau for m in self._musicien_morceaux[musicien] if m != morceau):
                charge_jour += 1
            
            if charge_jour >= self.max_load:
                conflicts += self.load_penalty * (charge_jour - self.max_load + 1)
        
        # 4. Bonus pour groupements (réduction des conflits)
        bonus = self._calculate_grouping_bonus(morceau, creneau)
        conflicts = max(0, conflicts - bonus)
        
        self._conflict_cache[cache_key] = conflicts
        return conflicts
    
    def _get_daily_load(self, musicien: str, jour: str, exclude_morceau: str = None) -> int:
        """Calcule la charge quotidienne d'un musicien."""
        charge = 0
        for morceau in self._musicien_morceaux[musicien]:
            if morceau == exclude_morceau:
                continue
            creneau = self.assignment.get(morceau)
            if creneau and creneau.startswith(jour + "_"):
                charge += 1
        return charge
    
    def _calculate_grouping_bonus(self, morceau: str, creneau: str) -> int:
        """Calcule le bonus de groupement pour un morceau/créneau."""
        bonus = 0
        jour = creneau.split("_")[0]
        slots_jour = self.creneaux_par_jour[jour]
        
        if creneau not in slots_jour:
            return 0
            
        creneau_idx = slots_jour.index(creneau)
        
        for musicien in self.repartition[morceau]:
            # Vérifier les créneaux adjacents
            for offset in [-1, 1]:
                adj_idx = creneau_idx + offset
                if 0 <= adj_idx < len(slots_jour):
                    adj_slot = slots_jour[adj_idx]
                    
                    # Vérifier si le musicien a une répétition dans le créneau adjacent
                    for autre_morceau in self._musicien_morceaux[musicien]:
                        if autre_morceau != morceau and self.assignment.get(autre_morceau) == adj_slot:
                            bonus += self.group_bonus
        
        return bonus
    
    def initialize_assignment(self):
        """Initialise l'assignation avec une solution heuristique."""
        print("Initialisation de l'assignation...")
        
        # Tri des morceaux par contraintes (plus contraints en premier)
        morceaux_tries = sorted(self.morceaux, 
                               key=lambda m: len(self.repartition[m]), 
                               reverse=True)
        
        for morceau in morceaux_tries:
            best_creneau = None
            min_conflicts = float('inf')
            
            # Essayer tous les créneaux
            for creneau in self.creneaux:
                conflicts = self.calculate_conflicts(morceau, creneau)
                if conflicts < min_conflicts:
                    min_conflicts = conflicts
                    best_creneau = creneau
            
            # Assigner si acceptable
            if best_creneau and min_conflicts < 1000:
                self.assignment[morceau] = best_creneau
            
        self._update_conflicts()
    
    def _update_conflicts(self):
        """Met à jour le compteur de conflits pour tous les morceaux."""
        self._conflict_cache.clear()
        self.conflicts = {}
        
        for morceau in self.morceaux:
            creneau = self.assignment.get(morceau)
            if creneau:
                self.conflicts[morceau] = self.calculate_conflicts(morceau, creneau)
            else:
                self.conflicts[morceau] = 1000  # Pénalité pour non-assignation
    
    def min_conflicts_step(self) -> bool:
        """Effectue une étape de l'algorithme min-conflicts."""
        # Trouver les morceaux avec des conflits
        morceaux_conflits = [(m, c) for m, c in self.conflicts.items() if c > 0]
        if not morceaux_conflits:
            return True  # Solution parfaite trouvée
        
        # Sélection avec biais vers les plus gros conflits
        morceaux_conflits.sort(key=lambda x: x[1], reverse=True)
        
        # most conflicted variable
        top_conflicted = morceaux_conflits[:min(3, len(morceaux_conflits))]
        morceau = random.choice([m for m, _ in top_conflicted])
        
        # Trouver le meilleur créneau pour ce morceau
        best_creneau = None
        min_conflicts = float('inf')
        
        # Essayer tous les créneaux + None (non-assigné)
        options = self.creneaux + [None]
        
        for creneau in options:
            old_creneau = self.assignment[morceau]
            self.assignment[morceau] = creneau
            if creneau:
                conflicts = self.calculate_conflicts(morceau, creneau)
            else:
                conflicts = 500  # pénalitépour non-assignation
            
            if conflicts < min_conflicts:
                min_conflicts = conflicts
                best_creneau = creneau
            
            # Restaurer l'état
            self.assignment[morceau] = old_creneau
        
        if best_creneau != self.assignment[morceau]:
            self.assignment[morceau] = best_creneau
            self._update_conflicts()
        
        return False
    
    def solve(self):
        """Résout le CSP avec l'algorithme min-conflicts (interface compatible)."""
        print("Début de la résolution avec min-conflicts...")
        start_time = time.time()
        
        best_solution = None
        best_cost = float('inf')
        
        for restart in range(self.max_restarts):
            print(f"Restart {restart + 1}/{self.max_restarts}")
            
            # Vérifier la limite de temps
            if time.time() - start_time > self.generation_time_limit:
                print("Limite de temps atteinte")
                break
        
            self.initialize_assignment()
            
            # Algorithme min-conflicts
            for iteration in range(self.max_iterations):
                if time.time() - start_time > self.generation_time_limit:
                    break
                
                if self.min_conflicts_step():
                    print(f"Solution parfaite trouvée en {iteration} itérations!")
                    self.status = "OPTIMAL"
                    self._finalize_solution()
                    return
                            
            # Sauvegarder la meilleure solution de ce restart
            current_cost = self._calculate_total_cost()
            if current_cost < best_cost:
                best_cost = current_cost
                best_solution = dict(self.assignment)
        
        # Utiliser la meilleure solution trouvée
        if best_solution:
            self.assignment = best_solution
            self.status = "FEASIBLE"
            print(f"Meilleure solution trouvée avec un coût de {best_cost}")
        else:
            self.status = "INFEASIBLE"
            print("Aucune solution trouvée")
        
        self._finalize_solution()
        
        # Affichage finale
        duration = time.time() - start_time
        assigned = sum(1 for v in self.assignment.values() if v is not None)
        total_conflicts = sum(self.conflicts.values())
        
        print(f"✅ Résolution terminée en {duration:.1f}s")
        print(f"✅ {assigned} morceaux assignés sur {len(self.morceaux)}")
        print(f"✅ Conflits totaux: {total_conflicts}")
    
    def _calculate_total_cost(self) -> int:
        """Calcule le coût total de la solution actuelle."""
        total = 0
        for morceau, creneau in self.assignment.items():
            if creneau:
                total += self.calculate_conflicts(morceau, creneau)
            else:
                total += 1000  # Pénalité pour non-assignation
        return total
    
    def _finalize_solution(self):
        """Finalise la solution et met à jour les structures compatibles."""
        self.solution = {}
        self.musiciens_absents_force.clear()
        
        for morceau, creneau in self.assignment.items():
            if creneau:
                self.solution[morceau] = creneau
                
                # Identifier les musiciens absents forcés
                for musicien in self.repartition[morceau]:
                    dispo = self.disponibilites.get(musicien, {}).get(creneau, "non")
                    if dispo == "non":
                        self.musiciens_absents_force[morceau].add(musicien)
    
    def generer_planning(self):
        """Interface compatible avec l'ancien code."""
        self.load_data()
        self.build_model()
        self.solve()


    def export_planning(self, directory=".", base_filename="planning"):
        import os
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        os.makedirs(directory, exist_ok=True)
        filename = f"{base_filename}_maybe{self.maybe_penalty}_load{self.max_load}_abs{self.seuil_absence}_timeout{self.generation_time_limit}.xlsx"
        path = os.path.join(directory, filename)
        # Constantes pour le tri et formatage (comme dans get_json_data)
        DAY_ORDER = {"Lundi":1, "Mardi":2, "Mercredi":3, "Jeudi":4,"Vendredi":5,"Samedi":6,"Dimanche":7, 
                    "LUN":1, "MAR":2, "MER":3, "JEU":4, "VEN":5, "SAM":6, "DIM":7}
        DAY_NAMES = {"LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi", "JEU":"Jeudi", 
                    "VEN":"Vendredi", "SAM":"Samedi", "DIM":"Dimanche"}

        def slot_sort_key(slot):
            jour, heures = format_slot(slot)
            reel_jour = jour.split(" ")[0]  # "Lundi 05" → "Lundi"
            return (DAY_ORDER[reel_jour], heures)

        def format_slot(slot: str) -> tuple[str,str]:
            """
            slot est au format "JOUR_DATE_START-END", ex. "LUN_05_14:00-16:00"
            Retourne (affichage_jour, affichage_heures),
            par exemple ("Lundi 05", "14:00-16:00").
            """
            try:
                jour_code, date_num, periode = slot.split("_")
            except ValueError:
                return slot, ""
            jour_aff = DAY_NAMES.get(jour_code, jour_code)
            return f"{jour_aff} {date_num}", periode

        # --- 1) DataFrame "Planning" ---
        planning_rows = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                planning_rows.append({
                    "Morceau": morceau,
                    "Jour": "Non assigné",
                    "Heures": "—",
                    "Participants": ", ".join(self.repartition.get(morceau, [])),
                    })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                planning_rows.append({
                    "Morceau": morceau,
                    "Jour": jour,
                    "Heures": heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
        df_planning = pd.DataFrame(planning_rows)
        
        # Tri du planning par jour
        if not df_planning.empty:
            df_planning["jour_order"] = df_planning["Jour"].apply(
                lambda x: DAY_ORDER.get(x.split(" ")[0], 99) if x != "Non assigné" else 100
            )
            df_planning.sort_values(["jour_order", "Heures"], inplace=True)
            df_planning.drop("jour_order", axis=1, inplace=True)

        # --- 2) DataFrames par semaine ---
        musiciens = sorted(self.musiciens)
        dispo_dfs = {}
        repart_dfs = {}

        # Pour chaque semaine détectée
        for w in self.weeks:
            # Sélection et tri des slots de la semaine w
            week_slots = [s for s in self.creneaux
                        if self.date2week[int(s.split("_")[1])] == w]
            week_slots = sorted(week_slots, key=slot_sort_key)

            # --- DataFrame Disponibilités pour cette semaine ---
            dispo_rows = []
            for slot in week_slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    row[m] = self.disponibilites.get(m, {}).get(slot, "non")
                dispo_rows.append(row)
            
            dispo_dfs[f"Dispo_Semaine_{w}"] = pd.DataFrame(dispo_rows)

            # --- DataFrame Répartition pour cette semaine ---
            repart_rows = []
            for slot in week_slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s==slot), None)
                row = {"Jour": jour, "Heures": heures, "Morceau": piece or ""}
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = self.disponibilites.get(m,{}).get(slot,"non")
                        row[m] = "Répète" if dispo=="oui" else "Répète"
                    else:
                        row[m] = ""
                repart_rows.append(row)
            
            repart_dfs[f"Repart_Semaine_{w}"] = pd.DataFrame(repart_rows)

        # --- 3) Écriture initiale via pandas ---
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_planning.to_excel(writer, sheet_name="Planning", index=False)
            
            # Écriture des onglets par semaine
            for sheet_name, df in dispo_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            for sheet_name, df in repart_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # --- 4) Re-ouverture et stylage avec openpyxl ---
        wb = load_workbook(filename)
        
        # Styles pour les disponibilités
        fill_yes    = PatternFill(fill_type="solid", fgColor="C6EFCE")  # Vert
        fill_maybe  = PatternFill(fill_type="solid", fgColor="FFEB9C")  # Jaune
        fill_no     = PatternFill(fill_type="solid", fgColor="F2DCDB")  # Rouge
        
        # Styles pour les répartitions
        fill_repete = PatternFill(fill_type="solid", fgColor="C6EFCE")  # Vert
        fill_absent = PatternFill(fill_type="solid", fgColor="FFB6C1")  # Rose
        fill_vide   = PatternFill(fill_type="solid", fgColor="D3D3D3")  # Gris
        
        # Stylage des onglets de disponibilités
        for sheet_name in dispo_dfs.keys():
            ws = wb[sheet_name]
            # Colonnes des musiciens commencent à la 3ème colonne (C)
            for row in ws.iter_rows(min_row=2, min_col=3):
                for cell in row:
                    v = (cell.value or "").strip().lower()
                    if v == "oui":
                        cell.fill = fill_yes
                    elif v in ("peut-être", "peut‐être"):
                        cell.fill = fill_maybe
                    else:
                        cell.fill = fill_no
        
        # Stylage des onglets de répartition
        for sheet_name in repart_dfs.keys():
            ws = wb[sheet_name]
            # Colonnes des musiciens commencent à la 4ème colonne (D)
            for row in ws.iter_rows(min_row=2, min_col=4):
                for cell in row:
                    v = (cell.value or "").strip()
                    if v == "Répète":
                        cell.fill = fill_repete
                    elif v == "Absent":
                        cell.fill = fill_absent
                    else:
                        cell.fill = fill_vide


        # --- Feuille Paramètres ---
        ws_params = wb.create_sheet(title="Paramètres")
        params_data = [
            ["Paramètre", "Valeur", f"Type de solution: {self.status}"],
            ["Fichier répartitions", self.repartitions_file],
            ["Fichier disponibilités", self.disponibilites_file],
            ["Pénalité maybe", self.maybe_penalty],
            ["Charge max", self.max_load],
            ["Pénalité charge", self.load_penalty],
            ["Bonus groupe", self.group_bonus],
            ["Mode absence", self.mode_absence],
            ["Seuil absence", self.seuil_absence],
            ["Temps limite génération", self.generation_time_limit]

        ]

        for row in params_data:
            ws_params.append(row)

        wb.save(path)
        return path

    def get_json_data(self):
        # pour trier et formater
        DAY_ORDER = {"Lundi":1, "Mardi":2, "Mercredi":3, "Jeudi":4,"Vendredi":5,"Samedi":6,"Dimanche":7, "LUN":1, "MAR":2, "MER":3, "JEU":4, "VEN":5, "SAM":6, "DIM":7}
        DAY_NAMES = {"LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi", "JEU":"Jeudi", "VEN":"Vendredi", "SAM":"Samedi", "DIM":"Dimanche"}

        def slot_sort_key(slot):
            jour, heures = format_slot(slot)
            reel_jour = jour.split(" ")[0]  # "Lundi 05" → "Lundi"
            return (DAY_ORDER[reel_jour], heures)

        def format_slot(slot: str) -> tuple[str,str]:
            """
            slot est au format "JOUR_DATE_START-END", ex. "LUN_05_14:00-16:00"
            Retourne (affichage_jour, affichage_heures),
            par exemple ("Lundi 05", "14:00-16:00").
            """
            # 1) découpage
            try:
                jour_code, date_num, periode = slot.split("_")
            except ValueError:
                # fallback si un slot mal formé traine
                return slot, ""
            # 2) optionnel : mappez "LUN" → "Lundi"
            jour_aff = DAY_NAMES.get(jour_code, jour_code)
            # 3) compose la chaîne complète pour l’entête "Jour"
            #    ici on peut afficher jour_aff + " " + date_num
            return f"{jour_aff} {date_num}", periode

        # 1) planning final (inchangé)
        planning = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                planning.append({
                    "Morceau": morceau,
                    "Jour":     "Non assigné",
                    "Heures":   "—",
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                planning.append({
                    "Morceau": morceau,
                    "Jour":     jour,
                    "Heures":   heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })

        # 2) DISPO & REPART pour **chaque** semaine détectée
        musiciens = sorted(self.musiciens)
        dispo_output   = {}
        repart_output  = {}

        # self.weeks contient déjà [1,2,3,…]
        for w in self.weeks:
            # sélection et tri des slots de la semaine w
            week_slots = [s for s in self.creneaux
                        if self.date2week[int(s.split("_")[1])] == w]
            week_slots = sorted(week_slots, key=slot_sort_key)

            dispo_rows  = []
            repart_rows = []

            # dispo
            for slot in week_slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    row[m] = self.disponibilites.get(m, {}).get(slot, "non")
                dispo_rows.append(row)

            # répartition
            for slot in week_slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s==slot), None)
                row = {"Jour": jour, "Heures": heures, "Morceau": piece or ""}
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = self.disponibilites.get(m,{}).get(slot,"non")
                        if dispo=="oui":
                            row[m] = "repete"
                        elif dispo =="non":
                            row[m] = "absent"
                        elif dispo == "peut-être":
                            row[m] = "maybe_absent"
                    else:
                        row[m] = "non"
                repart_rows.append(row)

            key = f"SEMAINE_{w}"
            dispo_output[key]  = dispo_rows
            repart_output[key] = repart_rows

        return {
            "planning":       planning,
            "disponibilites": dispo_output,
            "repartition":     repart_output
        }
    
  