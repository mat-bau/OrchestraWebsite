import pandas as pd
from ortools.sat.python import cp_model
import re
from collections import defaultdict
import time

class RepetitionScheduler:
    def __init__(self,
                 repartitions_file: str,
                 disponibilites_file: str,
                 maybe_penalty: int,
                 max_load: int,
                 load_penalty: int,
                 group_bonus: int,
                 mode_absence: str = "strict",
                 seuil_absence: int = 0):
        """
        Args:
            repartitions_file: Fichier Excel des répartitions donc avec les morceaux et participants
            disponibilites_file: Fichier Excel (avec Cally normalement) avec les disponibilités de chacun
        """
        self.repartitions_df = pd.read_excel(repartitions_file)
        self.disponibilites_df = pd.read_excel(disponibilites_file)

        self.musiciens = set()          # Liste des musiciens : ["Adèle", "Antoine", "Bastien...lol...bonhomme qui sourit 
                                        # à pleine dents"]
        self.musiciens_absents_force = defaultdict(set)  # morceau -> {musiciens absents mais contraints}
        self.absent_participants = defaultdict(lambda: defaultdict(list))  # morceau -> {musiciens absents mais assignés malgré leur indisponibilité}

        self.morceaux = []              # Liste des morceaux : ["morceau1" ,"morceau2"]
        self.repartition = {}
        self.repartition_absents = defaultdict(list)  # morceau -> [BoolVars des absences]

        self.disponibilites = {}        # musicien -> {créneau1: "oui"/"non"/"peut-être"} (ici j'utilise de l'optimisation
                                        # notamment pcq on peut avoir des "peut-être")
        
        self.creneaux = []              # Liste des créneaux : ["V_2_8-10", "S_1_14-16"]
        self.slot_index = {}            # transfo des créneaux en index {"V_1_8-10": 0, "V_1_10-14": 1, ...}  
        self.creneaux_par_jour = defaultdict(list)  # Créneaux par jour : {"LUN_0": ["LUN_0_8-10", "LUN_0_10-12"], ...}
        
        # Modele COP
        self.model = cp_model.CpModel()
        self.solver = cp_model.CpSolver()

        # Variables de décision
        self.assignments = {}  # morceau -> index du créneau

        self.penalties = [] # Liste des pénalités en foncton des critères d'optimisation
        self.solution = {}

        # Parmètre de pénalités
        self.maybe_penalty = maybe_penalty   # Pénalité pour "peut-être"
        self.max_load = max_load             # Nombre max de créneaux par jour
        self.load_penalty = load_penalty     # Pénalité si le musicien est surchargé
        self.group_bonus = group_bonus       # Gain pour les répétitions groupées
        self.mode_absence   = mode_absence      # Mode de gestion des absences ("strict", "flexible" ou "auto")
        self.seuil_absence = seuil_absence
        self.T             = None

    def transformer_simple(self, texte):
        """
        Fonction outil pour transformer le texte d'un créneau en format plus simple
        Args:
            texte: Le texte du créneau à transformer
        """
        import re
        texte_clean = texte.strip().replace('\n', ' ').replace('\r', ' ')
        match = re.search(r'(\w+\.)\s+(\d+).*?(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})', texte_clean)
        # print(f"Debug: Matching text '{texte_clean}' -> {match}")
        if not match:
            return None
        
        jour, date, h1, m1, h2, m2 = match.groups()

        jours = {'lun.': 'LUN', 'mar.': 'MAR', 'mer.': 'MER', 'jeu.': 'JEU',
                'ven.': 'VEN', 'sam.': 'SAM', 'dim.': 'DIM'}
        
        jour_conv = jours.get(jour, jour.upper())
        
        # a changer vous même en fonction de vos jours !!
        date_num = int(date)
        if date_num == 11 or date_num == 12 or date_num == 13:
            weekend = 1
        elif date_num == 18 or date_num == 19 or date_num == 20:
            weekend = 2
        elif date_num == 25 or date_num == 26 or date_num == 27:
            weekend = 3
        else:
            weekend = 0
        return f"{jour_conv}_{weekend}_{int(h1)}-{int(h2)}"
    
    def load_data(self):
        # A retravailler en fonction de la structure des fichiers d'input (je me base sur l'Excel 2024-2025)
        
        # 1. {Morceau: {Musicien1, Musicien2, ...}}
        instrument_cols = self.repartitions_df.columns[6:]
        for index, row in self.repartitions_df.iterrows():
            morceau = row['Titre']
            if pd.isna(morceau) or not any([not pd.isna(row[col]) for col in instrument_cols]): # lignes vides ou catégories (lignes sans musiciens)
                continue
            self.morceaux.append(morceau)
            self.repartition[morceau] = set()
            for col in instrument_cols:
                cellule = row[col]
                if pd.isna(cellule):
                    continue
                # plusieurs noms dans une cellule (peuvent être séparés par des virgules ou espaces)
                noms = [nom.strip() for nom in str(cellule).split(',')]
                for nom in noms:
                    if nom:
                        self.musiciens.add(nom)
                        self.repartition[morceau].add(nom)
        
        # 2. {Musicien: {Créneau1: "oui"/"non"/"peut-être"}}
        dispo_cols = self.disponibilites_df.columns[2:]  # On ignore les deux premières colonnes
        for index, row in self.disponibilites_df.iterrows():
            musicien = str(row['Nom']).strip().title()  
            if pd.isna(musicien):
                continue
            local_dispo = {}
            for col in dispo_cols:
                creneau = self.transformer_simple(str(col).strip())
                local_dispo[creneau] = str(row[col]).strip().lower() if not pd.isna(row[col]) else "non"
            self.disponibilites[musicien] = local_dispo
        
        # 3. Créneaux
        self.creneaux = list(next(iter(self.disponibilites.values())).keys())
        self.creneaux.sort()
        self.slot_index = {slot: i for i, slot in enumerate(self.creneaux)}
        for creaneau in self.creneaux:
            jour = "_".join(creaneau.split("_")[:2])
            self.creneaux_par_jour[jour].append(creaneau)


    def define_variables(self):
        """
        - self.is_assigned[morceau] = BoolVar: 1 si on affecte le morceau à un slot
        - self.assignments[morceau] = IntVar en [0..N], N == len(self.creneaux) => 
            valeurs 0..N-1 = index de créneau, N = “pas affecté”
        """
        self.is_assigned = {}
        self.assignments = {}

        domain_max = len(self.creneaux)  # valeur hors plage = “pas assigné”
        for morceau in self.morceaux:
            # Booléen d'affectation
            b = self.model.NewBoolVar(f"is_assigned_{morceau}")
            self.is_assigned[morceau] = b

            # IntVar slot (0..N)
            v = self.model.NewIntVar(0, domain_max, f"assignment_{morceau}")
            self.assignments[morceau] = v

            # si b=1 alors v < domain_max ; si b=0 alors v == domain_max
            self.model.Add(v < domain_max).OnlyEnforceIf(b)
            self.model.Add(v == domain_max).OnlyEnforceIf(b.Not())



    # 1st constraint : dipsonibilité "oui" "non" "peut-être"
    def add_disponibility_constraints(self):
        # si mode == "auto", on crée T = maxi d’absents autorisés sur un slot
        if self.mode_absence == "auto":
            # bornes 0…(nombre total de musiciens)
            total_mus = sum(len(mus) for mus in self.repartition.values())
            self.T = self.model.NewIntVar(0, total_mus, "T_max_abs")

        if self.mode_absence == "fixed" and self.seuil_absence == 0: # ancienne version (qui est dans la main branch) mais en gros prend les peut-etre un peu moins stricte jsp pq faut que j'y regarde
            for morceau, musiciens in self.repartition.items():
                for musicien in musiciens:
                    for slot in self.creneaux:
                        slot_idx = self.slot_index[slot]
                        is_dispo = self.disponibilites[musicien].get(slot, "non").lower()

                        is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_here")
                        self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                        self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())

                        # Si dispo == non -> contrainte dure (seulement si le morceau est assigné)
                        if is_dispo == "non":
                            self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(self.is_assigned[morceau])

                        # Si dispo == peut-être -> pénalité
                        if is_dispo == "peut-être":
                            penalty = self.model.NewIntVar(0, self.maybe_penalty, f"penalty_{morceau}_{slot}_{musicien}")
                            self.model.Add(penalty == self.maybe_penalty).OnlyEnforceIf(is_here)
                            self.model.Add(penalty == 0).OnlyEnforceIf(is_here.Not())
                            self.penalties.append(penalty)
        else:
            for morceau, musiciens in self.repartition.items():
                for slot in self.creneaux:
                    slot_idx = self.slot_index[slot]
                    is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_here")
                    self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                    self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())
                    
                    absent_flags = []
                    for musicien in musiciens:
                        dispo = self.disponibilites[musicien].get(slot,"non").lower()
                        if dispo == "oui":
                            continue
                        absent = self.model.NewBoolVar(f"{morceau}_{slot}_{musicien}_absent")
                        self.model.Add(is_here == 1).OnlyEnforceIf(absent)
                        self.model.Add(is_here == 0).OnlyEnforceIf(absent.Not())
                        absent_flags.append(absent)
                        # pénalités "non" / "peut-être"
                        if dispo == "non":
                            pen = self.model.NewIntVar(0,1, f"pen_non_{morceau}_{slot}_{musicien}")
                            self.model.Add(pen == 1).OnlyEnforceIf(absent)
                            self.model.Add(pen == 0).OnlyEnforceIf(absent.Not())
                        else:
                            pen = self.model.NewIntVar(0,self.maybe_penalty,
                                                        f"pen_maybe_{morceau}_{slot}_{musicien}")
                            self.model.Add(pen == self.maybe_penalty).OnlyEnforceIf(absent)
                            self.model.Add(pen == 0).OnlyEnforceIf(absent.Not())
                            self.penalties.append(pen)
                            self.absent_participants.setdefault(morceau, {})\
                                                    .setdefault(slot, []).append((musicien, absent))

                    if not absent_flags:
                        continue
                    if self.mode_absence == "strict":
                        self.model.Add(sum(absent_flags) == 0).OnlyEnforceIf(is_here)

                    elif self.mode_absence == "fixed":
                        sum_abs = self.model.NewIntVar(0, len(absent_flags),
                                                    f"{morceau}_{slot}_sum_abs")
                        self.model.Add(sum_abs == sum(absent_flags)).OnlyEnforceIf(is_here)
                        self.model.Add(sum_abs <= self.seuil_absence).OnlyEnforceIf(is_here)

                    else:
                        sum_abs = self.model.NewIntVar(0, len(absent_flags),
                                                    f"{morceau}_{slot}_sum_abs")
                        self.model.Add(sum_abs == sum(absent_flags)).OnlyEnforceIf(is_here)
                        self.model.Add(sum_abs <= self.T).OnlyEnforceIf(is_here)


    # 2nd: Un créneau n'accueille qu'un morceau
    def add_slot_constraints(self):
        for slot in self.creneaux:
            slot_idx = self.slot_index[slot]
            is_assigned = []
            for morceau in self.morceaux:
                is_at_slot = self.model.NewBoolVar(f"{morceau}_is_at_{slot}")
                self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_at_slot)
                self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_at_slot.Not())
                is_assigned.append(is_at_slot)
            self.model.Add(sum(is_assigned) <= 1)

    # 3rd: Eviter les journées trop chargées
    def add_daily_load_constraints(self):
        for musicien in self.musiciens:
            for jour, slots in self.creneaux_par_jour.items():
                morceaux_du_musicien = [m for m in self.morceaux if musicien in self.repartition[m]]
                occupied_slots = []
                for slot in slots:
                    present = self.model.NewBoolVar(f"{musicien}_present_{jour}_{slot}")
                    clauses = []
                    for morceau in morceaux_du_musicien:
                        is_at_slot = self.model.NewBoolVar(f"{morceau}_is_at_{slot}")
                        self.model.Add(self.assignments[morceau] == self.slot_index[slot]).OnlyEnforceIf(is_at_slot)
                        self.model.Add(self.assignments[morceau] != self.slot_index[slot]).OnlyEnforceIf(is_at_slot.Not())
                        clauses.append(is_at_slot)
                    if clauses:
                        self.model.AddMaxEquality(present,clauses)
                    else:
                        self.model.Add(present == 0)
                    occupied_slots.append(present)
                    nb_slots = self.model.NewIntVar(0, len(slots), f"nb_slots_{musicien}_{jour}")
                    self.model.Add(nb_slots == sum(occupied_slots))
                    # Pénalité si le musicien est surchargé
                    is_overloaded = self.model.NewBoolVar(f"{musicien}_overloaded_{jour}")
                    self.model.Add(nb_slots > self.max_load).OnlyEnforceIf(is_overloaded)
                    self.model.Add(nb_slots <= self.max_load).OnlyEnforceIf(is_overloaded.Not())
                    penalty = self.model.NewIntVar(0, self.load_penalty, f"penalty_{musicien}_{jour}")
                    self.model.Add(penalty == self.load_penalty).OnlyEnforceIf(is_overloaded)
                    self.model.Add(penalty == 0).OnlyEnforceIf(is_overloaded.Not())
                    self.penalties.append(penalty)

    # 4th: Pénalités pour les répétitions groupées
    def add_penalites_repetitions_groupees(self):
        for musicien in self.musiciens:
            for jour, slots in self.creneaux_par_jour.items():
                morceaux_joues = [m for m in self.morceaux if musicien in self.repartition[m]]
                
                presence = []
                for slot in slots:
                    slot_idx = self.slot_index[slot]
                    var = self.model.NewBoolVar(f"{musicien}_{slot}_present")
                    
                    clauses = []
                    for morceau in morceaux_joues:
                        is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_at")
                        self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                        self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())
                        clauses.append(is_here)
                    
                    if clauses:
                        self.model.AddMaxEquality(var, clauses)
                    else:
                        self.model.Add(var == 0)
                    
                    presence.append(var)
                for i in range(len(presence) - 1):
                    bloc = self.model.NewBoolVar(f"{musicien}_{jour}_bloc_{i}")
                    self.model.AddBoolAnd([presence[i], presence[i + 1]]).OnlyEnforceIf(bloc)
                    self.model.AddBoolOr([presence[i].Not(), presence[i + 1].Not()]).OnlyEnforceIf(bloc.Not())

                    reward = self.model.NewIntVar(-self.group_bonus, 0, f"reward_bloc_{musicien}_{jour}_{i}")
                    self.model.Add(reward == -self.group_bonus).OnlyEnforceIf(bloc)
                    self.model.Add(reward == 0).OnlyEnforceIf(bloc.Not())
                    self.penalties.append(reward)  

    def define_objective(self):
        penalty_not_assigned_weight = 10000
        # 1) penalty pour non‐assignés
        for morceau in self.morceaux:
            p = self.model.NewIntVar(0, penalty_not_assigned_weight,
                                     f"penalite_non_assigne_{morceau}")
            self.model.Add(p == penalty_not_assigned_weight).OnlyEnforceIf(self.is_assigned[morceau].Not())
            self.model.Add(p == 0).OnlyEnforceIf(self.is_assigned[morceau])
            self.penalties.append(p)

        # 2) si mode "auto", on minimise ensuite T
        objective = sum(self.penalties)
        if self.mode_absence == "auto":
            W2 = 10000
            objective = objective + self.T * W2

        # 3) on inclut toutes les pénalités "non"/"peut-être"
        self.model.Minimize(objective)

    def build_model(self):
        # 1) (re)création du modèle
        self.model = cp_model.CpModel()
        # 2) définir les variables
        self.define_variables()

        # 3) ajouter toutes les contraintes
        self.add_disponibility_constraints()
        self.add_slot_constraints()
        self.add_daily_load_constraints()
        self.add_penalites_repetitions_groupees()

        # 4) définir l'objectif
        self.define_objective()


    def solve(self,
            time_limit_sec: float = 120,
            num_workers: int = 12):
        """
        Construit le modèle (via build_model), résout, et extrait :
        - status
        - solution dict morceau->slot
        - num_unassigned
        - total_penalty
        """

        self.solver = cp_model.CpSolver()
        self.solver.parameters.max_time_in_seconds = time_limit_sec
        self.solver.parameters.num_search_workers = num_workers

        self.penalties = []
        self.musiciens_absents_force.clear()
        self.build_model()

        # --- Solve ---
        start = time.time()
        status = self.solver.Solve(self.model)
        duration = time.time() - start
        print(f"Solve status = {self.solver.StatusName(status)} en {duration:.1f}s")

        # --- Post‐processing ---
        num_unassigned = 0
        total_penalty = 0

        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            # récupérer la pénalité
            total_penalty = sum(self.solver.Value(p) for p in self.penalties)

            for morceau in self.morceaux:
                if self.solver.Value(self.is_assigned[morceau]):
                    idx = self.solver.Value(self.assignments[morceau])
                    slot = self.creneaux[idx]
                    self.solution[morceau] = slot

                    # repérer les absents forcés
                    for (mus, flag) in self.absent_participants[morceau].get(slot, []):
                        if self.solver.Value(flag):
                            self.musiciens_absents_force[morceau].add(mus)
                else:
                    num_unassigned += 1

            print(f"✅ {len(self.solution)} assignés, {num_unassigned} non-assignés, pénalité totale = {total_penalty}")
        else:
            print("⚠️ Aucune solution trouvée")

        return status, self.solution, num_unassigned, total_penalty


    def generer_planning(self):
        self.load_data()
        self.build_model()
        self.solve()

    def export_planning(self, output_file):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        # --- 1) DataFrame "Planning" ---
        rows = []
        for morceau in self.morceaux:
            if morceau in self.solution:
                slot = self.solution[morceau]
                jour, horaire = slot.split("_", 1)
            else:
                jour, horaire = "Non assigné", "Non assigné"
            participants = ", ".join(self.repartition.get(morceau, []))
            rows.append({
                "Morceau": morceau,
                "Jour":    jour,
                "Créneau": horaire,
                "Participants": participants
            })
        df_plan = pd.DataFrame(rows)
        # tri des jours
        ordre_jours = ["LUN","MAR","MER","JEU","VEN","SAM","DIM","Non assigné"]
        df_plan["Jour"] = pd.Categorical(df_plan["Jour"],
                                         categories=ordre_jours,
                                         ordered=True)
        df_plan.sort_values(["Jour","Créneau"], inplace=True)

        # --- 2) DataFrame "Disponibilités" ---
        # on veut un tableau à plat : une ligne par créneau, colonnes musiciens
        # tri des créneaux
        def jour_cle(s):
            j = s.split("_",1)[0]
            return (ordre_jours.index(j) if j in ordre_jours else 99, s)
        slots = sorted(self.creneaux, key=jour_cle)
        musiciens = sorted(self.musiciens)

        dispo_dict = {"Créneau": slots}
        for m in musiciens:
            dispo_dict[m] = [
                self.disponibilites.get(m, {})
                                   .get(s, "non")
                                   .strip().lower()
                for s in slots
            ]
        df_dispo = pd.DataFrame(dispo_dict)

        # --- 3) DataFrames "Répartition_W1" et "Répartition_W2" ---
        rep_dfs = {}
        for wk in (1,2):
            wk_slots = [s for s in slots if s.endswith(f"_{wk}")]
            rep_dict = {"Morceau": [], "Créneau": wk_slots}
            for m in musiciens:
                rep_dict[m] = []
            for s in wk_slots:
                # retrouver le morceau joué à ce créneau
                piece = next((
                    p for p,sl in self.solution.items() if sl==s
                ), None)
                rep_dict["Morceau"].append(piece or "Aucun")
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        rep_dict[m].append("Répète")
                    else:
                        rep_dict[m].append("")
            rep_dfs[f"Répartition_W{wk}"] = pd.DataFrame(rep_dict)

        # --- 4) Écriture initiale via pandas ---
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_plan.to_excel(writer,      sheet_name="Planning",      index=False)
            df_dispo.to_excel(writer,     sheet_name="Disponibilités", index=False)
            for sheet, df in rep_dfs.items():
                df.to_excel(writer, sheet_name=sheet, index=False)

        # --- 5) Re-ouverture et stylage avec openpyxl ---
        wb = load_workbook(output_file)
        # styles dispo
        fill_yes    = PatternFill(fill_type="solid", fgColor="C6EFCE")
        fill_maybe  = PatternFill(fill_type="solid", fgColor="FFEB9C")
        fill_no     = PatternFill(fill_type="solid", fgColor="F2DCDB")
        ws = wb["Disponibilités"]
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                v = (cell.value or "").strip().lower()
                if v=="oui":       cell.fill = fill_yes
                elif v in ("peut-être","peut‐être"): cell.fill = fill_maybe
                else:               cell.fill = fill_no

        # styles répartition
        fill_repete     = PatternFill(fill_type="solid", fgColor="C6EFCE")
        fill_non_repete = PatternFill(fill_type="solid", fgColor="D3D3D3")
        for wk in (1,2):
            ws = wb[f"Répartition_W{wk}"]
            # colonnes à partir de la 3ᵉ (les musiciens)
            for row in ws.iter_rows(min_row=2, min_col=3):
                for cell in row:
                    if cell.value=="Répète":
                        cell.fill = fill_repete
                    else:
                        cell.fill = fill_non_repete

        wb.save(output_file)
        print(f"✅ Planning exporté dans {output_file}")

    def get_json_data(self):
        """
        Construit les données attendues par le frontend sous forme de dictionnaire,
        en séparant les colonnes Jour / Heures et en triant correctement les créneaux.
        """
        # ordre des jours pour le tri
        DAY_ORDER = {"LUN":1, "MAR":2, "MER":3, "JEU":4,
                    "VEN":5, "SAM":6, "DIM":7}
        # noms complets pour l’affichage
        DAY_NAMES = {
            "LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi",
            "JEU":"Jeudi","VEN":"Vendredi",
            "SAM":"Samedi","DIM":"Dimanche"
        }

        def slot_sort_key(slot: str):
            # ex slot = "VEN_1_10-12"
            jour, _, plage = slot.split("_")
            start, _ = plage.split("-")
            h, m = (start.split(":") + ["00"])[:2]
            return (DAY_ORDER[jour], int(h), int(m))

        def format_slot(slot: str):
            # retourne (Jour, Heures)
            jour, _, plage = slot.split("_")
            start, end = plage.split("-")
            def fmt(x):
                h, m = (x.split(":") + ["00"])[:2]
                return f"{int(h)}h{m:>02}"
            return DAY_NAMES[jour], f"{fmt(start)}-{fmt(end)}"

        # 1) planning général Morceau / Jour / Heures
        data = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                data.append({
                    "Morceau": morceau,
                    "Jour":     "Non assigné",
                    "Heures":   "—",
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                data.append({
                    "Morceau": morceau,
                    "Jour":     jour,
                    "Heures":   heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })

        # 2) disponibilités et répartition par weekend
        disponibilites = {"weekend1": [], "weekend2": []}
        repartition     = {"weekend1": [], "weekend2": []}
        musiciens = sorted(self.musiciens)

        for w in (1, 2):
            raw_slots = [s for s in self.creneaux if f"_{w}_" in s]
            slots     = sorted(raw_slots, key=slot_sort_key)

            # tableau disponibilités
            for slot in slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    d = (self.disponibilites.get(m, {})
                                    .get(slot, "non")
                                    .strip().lower())
                    row[m] = d  # "oui"/"non"/"peut-être"
                disponibilites[f"weekend{w}"].append(row)

            # tableau répartition (avec morceau et état repete/absent/non)
            for slot in slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s == slot), None)
                row = {
                    "Morceau": piece or "",
                    "Jour":    jour,
                    "Heures":  heures
                }
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = (self.disponibilites.get(m, {})
                                        .get(slot, "non")
                                        .strip().lower())
                        row[m] = "repete" if dispo == "oui" else "absent"
                    else:
                        row[m] = "non"
                repartition[f"weekend{w}"].append(row)

        return {
            "planning":       data,
            "disponibilites": disponibilites,
            "repartition":     repartition
        }
