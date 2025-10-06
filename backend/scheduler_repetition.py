"""
Description : Planificateur de répétitions pour l'Orchestraaaaa
Licence : On devrait peut-être mettre une licence hein
Anno : 43
Auteur : Mateo Bauvir
"""

import pandas as pd
from ortools.sat.python import cp_model
import re
from collections import defaultdict
import time

class RepetitionScheduler:
    def __init__(self,
                 repartitions_file: str,
                 disponibilites_file: str,
                 maybe_penalty: int,
                 max_load: int,
                 load_penalty: int,
                 group_bonus: int,
                 mode_absence: str = "strict",
                 seuil_absence: int = 0,
                 generation_time_limit: int = 30):
        """
        Args:
            repartitions_file: Fichier Excel des répartitions donc avec les morceaux et participants
            disponibilites_file: Fichier Excel (avec Cally normalement) avec les disponibilités de chacun
        """
        self.repartitions_file = repartitions_file
        self.disponibilites_file = disponibilites_file
        self.repartitions_df = pd.read_excel(repartitions_file)
        self.disponibilites_df = pd.read_excel(disponibilites_file)

        self.musiciens = set()          # Liste des musiciens : ["Adèle", "Antoine", "Bastien...lol...bonhomme qui sourit à pleine dents"]
        self.musiciens_absents_force = defaultdict(set)  # morceau -> {musiciens absents mais contraints}
        self.absent_participants = defaultdict(lambda: defaultdict(list))  # morceau -> {musiciens absents mais assignés malgré leur indisponibilité}

        self.morceaux = []              # Liste des morceaux : ["morceau1" ,"morceau2"]
        self.repartition = {}
        self.repartition_absents = defaultdict(list)  # morceau -> [BoolVars des absences]

        self.disponibilites = {}        # musicien -> {créneau1: "oui"/"non"/"peut-être"} (ici j'utilise de l'optimisation
                                        # notamment pcq on peut avoir des "peut-être")
        
        self.creneaux = []              # Liste des créneaux : ["V_2_8-10", "S_1_14-16"]
        self.slot_index = {}            # transfo des créneaux en index {"V_1_8-10": 0, "V_1_10-14": 1, ...}  
        self.creneaux_par_jour = defaultdict(list)  # Créneaux par jour : {"LUN_0": ["LUN_0_8-10", "LUN_0_10-12"], ...}
        
        # Modele COP
        self.model = cp_model.CpModel()
        self.solver = cp_model.CpSolver()
        self.status = None  # Statut de la résolution

        # Variables de décision
        self.assignments = {}  # morceau -> index du créneau

        self.penalties = [] # Liste des pénalités en foncton des critères d'optimisation
        self.solution = {}

        # Parmètre de pénalités
        self.maybe_penalty = maybe_penalty   # Pénalité pour "peut-être"
        self.max_load = max_load             # Nombre max de créneaux par jour
        self.load_penalty = load_penalty     # Pénalité si le musicien est surchargé
        self.group_bonus = group_bonus       # Gain pour les répétitions groupées
        self.mode_absence   = mode_absence      # Mode de gestion des absences ("strict", "flexible" ou "auto")
        self.seuil_absence = seuil_absence
        self.T             = None
        self.generation_time_limit = generation_time_limit  # limite de temps laissé à la génération du planning

    def transformer_simple(self, texte: str) -> Optional[Dict]:
        """Transforme un texte de créneau en dictionnaire structuré."""
        t = texte.strip().replace("\n", " ").replace("\r", " ")
        m = re.search(r"(\w+\.)\s+(\d+).*?(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})", t)
        if not m:
            return None
        
        jour_txt, jour_chiffre, H1, M1, H2, M2 = m.groups()
        jours = {
            'lun.': 'LUN', 'mar.': 'MAR', 'mer.': 'MER',
            'jeu.': 'JEU', 'ven.': 'VEN', 'sam.': 'SAM', 'dim.': 'DIM'
        }
        jour = jours.get(jour_txt.lower(), jour_txt.upper().rstrip('.'))
        
        return {
            'jour': jour,
            'date': int(jour_chiffre),
            'h1': int(H1), 'm1': int(M1),
            'h2': int(H2), 'm2': int(M2)
        }
    
    def load_data(self):
        """Charge les données depuis les fichiers Excel."""
        # print("Chargement des données...")
        
        # Charger les répartitions si le fichier existe
        if hasattr(self, 'repartitions_file') and self.repartitions_file:
            self.repartitions_df = pd.read_excel(self.repartitions_file)
            instrument_cols = self.repartitions_df.columns[6:]
                    
            for _, row in self.repartitions_df.iterrows():
                morceau = row['Titre']
                if pd.isna(morceau):
                    continue
                                
                has_musicians = any(not pd.isna(row[c]) for c in instrument_cols)
                if not has_musicians:
                    continue
                                
                self.morceaux.append(morceau)
                self.repartition[morceau] = set()
                            
                for col in instrument_cols:
                    cellule = row[col]
                    if pd.isna(cellule):
                        continue
                                        
                    for nom in str(cellule).split(','):
                        nom = nom.strip()
                        if nom:
                            self.musiciens.add(nom)
                            self.repartition[morceau].add(nom)
                            self._musicien_morceaux[nom].append(morceau)
        
        # Charger les disponibilités
        self.disponibilites_df = pd.read_excel(self.disponibilites_file)
        all_dates = set()
        
        # Détecter c'est quel format
        if 'Nom' in self.disponibilites_df.columns:
            # ANCIEN FORMAT (que j'utilisais pour tester au début)
            dispo_cols = self.disponibilites_df.columns[2:]
            
            for _, row in self.disponibilites_df.iterrows():
                musicien = str(row['Nom']).strip().title()
                if pd.isna(musicien) or not musicien:
                    continue
                                
                self.disponibilites[musicien] = {}
                            
                for col in dispo_cols:
                    info = self.transformer_simple(str(col))
                    if not info:
                        continue
                                        
                    d = info['date']
                    j = info['jour']
                    h1, m1, h2, m2 = info['h1'], info['m1'], info['h2'], info['m2']
                    all_dates.add(d)
                    start = f"{h1:02d}:{m1:02d}"
                    end = f"{h2:02d}:{m2:02d}"
                    slot = f"{j}_{d:02d}_{start}-{end}"
                                    
                    val = str(row[col]).strip().lower() if not pd.isna(row[col]) else "no"
                    self.disponibilites[musicien][slot] = val
        else:
            # le format de cally
            for idx, row in self.disponibilites_df.iterrows():
                for col_idx, col in enumerate(self.disponibilites_df.columns[2:], start=2):  # Skip les 2 premières colonnes
                    cell_value = str(row[col]) if not pd.isna(row[col]) else ""
                    if cell_value and cell_value != "nan":
                        if any(day in cell_value for day in ['lun.', 'mar.', 'mer.', 'jeu.', 'ven.', 'sam.', 'dim.']):
                            clean_text = cell_value.replace('\n', ' ').replace('[93%]', '').strip()
                            info = self.transformer_simple(clean_text)
                            if info:
                                d = info['date']
                                j = info['jour']
                                h1, m1, h2, m2 = info['h1'], info['m1'], info['h2'], info['m2']
                                all_dates.add(d)
                                start = f"{h1:02d}:{m1:02d}"
                                end = f"{h2:02d}:{m2:02d}"
                                slot = f"{j}_{d:02d}_{start}-{end}"
                                self.creneaux.append(slot)
            
            # Deuxième étape: traiter les réponses des musiciens
            for idx, row in self.disponibilites_df.iterrows():
                if idx <= 3:  # Skip les lignes d'en-tête et de créneaux
                    continue
                    
                first_col = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
                
                # Vérifier si c'est une ligne de musicien (contient un email dans la 2ème colonne)
                if len(row) > 1 and not pd.isna(row.iloc[1]) and '@' in str(row.iloc[1]):
                    musicien = first_col.title()
                    if musicien and musicien != 'Nan':
                        self.musiciens.add(musicien)
                        self.disponibilites[musicien] = {}
                        
                        
                        # Parcourir les réponses (à partir de la colonne 2)
                        for col_idx in range(2, len(row)):
                            if col_idx - 2 < len(self.creneaux):  # Vérifier qu'on a un créneau correspondant
                                slot = self.creneaux[col_idx - 2]
                                val = str(row.iloc[col_idx]).strip().lower() if not pd.isna(row.iloc[col_idx]) else "no"
                                
                                if val in ['yes', 'oui']:
                                    self.disponibilites[musicien][slot] = "yes"
                                elif val in ['maybe', 'peut-être']:
                                    self.disponibilites[musicien][slot] = "maybe"  
                                else:
                                    self.disponibilites[musicien][slot] = "no"
            
            # Si pas de morceaux définis, créer un morceau par défaut pour le nouveau format
            if not self.morceaux and self.musiciens:
                morceau_default = "Session_Planning"
                self.morceaux.append(morceau_default)
                self.repartition[morceau_default] = self.musiciens.copy()
                
                for musicien in self.musiciens:
                    self._musicien_morceaux[musicien].append(morceau_default)
        
        # Finalisation commune aux deux formats
        if self.disponibilites:
            premier = next(iter(self.disponibilites.values()))
                            
            def _key(s):
                _, dd, plage = s.split("_")
                start, _ = plage.split("-")
                h, m = start.split(":")
                return (int(dd), int(h), int(m))
            
            # Pour l'ancien format, les créneaux sont extraits des disponibilités
            if 'Nom' in self.disponibilites_df.columns:
                self.creneaux = sorted(premier.keys(), key=_key)
            else:
                # Pour le nouveau format, les créneaux sont déjà extraits, on les trie juste
                self.creneaux = sorted(self.creneaux, key=_key)
                
            self.slot_index = {slot: i for i, slot in enumerate(self.creneaux)}
                            
            # Regroupement par jour
            for slot in self.creneaux:
                jour = slot.split("_")[0]
                self.creneaux_par_jour[jour].append(slot)
                    
            dates_sorted = sorted(all_dates)
            if dates_sorted:
                base = dates_sorted[0]
                self.date2week = {d: ((d - base) // 7) + 1 for d in dates_sorted}
                self.weeks = sorted(set(self.date2week.values()))
            else:
                self.weeks = []
                    


    def define_variables(self):
        """
        - self.is_assigned[morceau] = BoolVar: 1 si on affecte le morceau à un slot
        - self.assignments[morceau] = IntVar en [0..N], N == len(self.creneaux) => 
            valeurs 0..N-1 = index de créneau, N = “pas affecté”
        """
        self.is_assigned = {}
        self.assignments = {}

        domain_max = len(self.creneaux)  # valeur hors plage = “pas assigné”
        for morceau in self.morceaux:
            b = self.model.NewBoolVar(f"is_assigned_{morceau}")
            self.is_assigned[morceau] = b
            v = self.model.NewIntVar(0, domain_max, f"assignment_{morceau}")
            self.assignments[morceau] = v
            self.model.Add(v < domain_max).OnlyEnforceIf(b)
            self.model.Add(v == domain_max).OnlyEnforceIf(b.Not())



    # 1st constraint : dipsonibilité "oui" "non" "peut-être"
    def add_disponibility_constraints(self):
        # si mode == "auto", on crée T = maxi d’absents autorisés sur un slot
        if self.mode_absence == "auto":
            # bornes 0…(nombre total de musiciens)
            total_mus = sum(len(mus) for mus in self.repartition.values())
            self.T = self.model.NewIntVar(0, total_mus, "T_max_abs")

        if self.mode_absence == "fixed" and self.seuil_absence == 0: # ancienne version
            for morceau, musiciens in self.repartition.items():
                for musicien in musiciens:
                    for slot in self.creneaux:
                        slot_idx = self.slot_index[slot]
                        is_dispo = self.disponibilites[musicien].get(slot, "non").lower()

                        is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_here")
                        self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                        self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())

                        # Si dispo == non -> contrainte dure (seulement si le morceau est assigné)
                        if is_dispo == "non":
                            self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(self.is_assigned[morceau])

                        # Si dispo == peut-être -> pénalité
                        if is_dispo == "peut-être":
                            penalty = self.model.NewIntVar(0, self.maybe_penalty, f"penalty_{morceau}_{slot}_{musicien}")
                            self.model.Add(penalty == self.maybe_penalty).OnlyEnforceIf(is_here)
                            self.model.Add(penalty == 0).OnlyEnforceIf(is_here.Not())
                            self.penalties.append(penalty)
        else:
            for morceau, musiciens in self.repartition.items():
                for slot in self.creneaux:
                    slot_idx = self.slot_index[slot]
                    is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_here")
                    self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                    self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())
                    
                    absent_flags = []
                    for musicien in musiciens:
                        dispo = self.disponibilites[musicien].get(slot,"non").lower()
                        if dispo == "oui":
                            continue
                        absent = self.model.NewBoolVar(f"{morceau}_{slot}_{musicien}_absent")
                        self.model.Add(is_here == 1).OnlyEnforceIf(absent)
                        self.model.Add(is_here == 0).OnlyEnforceIf(absent.Not())
                        absent_flags.append(absent)
                        # pénalités "non" / "peut-être"
                        if dispo == "non":
                            pen = self.model.NewIntVar(0,1, f"pen_non_{morceau}_{slot}_{musicien}")
                            self.model.Add(pen == 1).OnlyEnforceIf(absent)
                            self.model.Add(pen == 0).OnlyEnforceIf(absent.Not())
                        else:
                            pen = self.model.NewIntVar(0,self.maybe_penalty,
                                                        f"pen_maybe_{morceau}_{slot}_{musicien}")
                            self.model.Add(pen == self.maybe_penalty).OnlyEnforceIf(absent)
                            self.model.Add(pen == 0).OnlyEnforceIf(absent.Not())
                            self.penalties.append(pen)
                            self.absent_participants.setdefault(morceau, {})\
                                                    .setdefault(slot, []).append((musicien, absent))

                    if not absent_flags:
                        continue

                    elif self.mode_absence == "fixed":
                        sum_abs = self.model.NewIntVar(0, len(absent_flags),
                                                    f"{morceau}_{slot}_sum_abs")
                        self.model.Add(sum_abs == sum(absent_flags)).OnlyEnforceIf(is_here)
                        self.model.Add(sum_abs <= self.seuil_absence).OnlyEnforceIf(is_here)

                    else:
                        sum_abs = self.model.NewIntVar(0, len(absent_flags),
                                                    f"{morceau}_{slot}_sum_abs")
                        self.model.Add(sum_abs == sum(absent_flags)).OnlyEnforceIf(is_here)
                        self.model.Add(sum_abs <= self.T).OnlyEnforceIf(is_here)


    # 2nd: Un créneau n'accueille qu'un morceau
    def add_slot_constraints(self):
        for slot in self.creneaux:
            slot_idx = self.slot_index[slot]
            is_assigned = []
            for morceau in self.morceaux:
                is_at_slot = self.model.NewBoolVar(f"{morceau}_is_at_{slot}")
                self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_at_slot)
                self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_at_slot.Not())
                is_assigned.append(is_at_slot)
            self.model.Add(sum(is_assigned) <= 1)

    # 3rd: Eviter les journées trop chargées
    def add_daily_load_constraints(self):
        for musicien in self.musiciens:
            for jour, slots in self.creneaux_par_jour.items():
                morceaux_du_musicien = [m for m in self.morceaux if musicien in self.repartition[m]]
                occupied_slots = []
                for slot in slots:
                    present = self.model.NewBoolVar(f"{musicien}_present_{jour}_{slot}")
                    clauses = []
                    for morceau in morceaux_du_musicien:
                        is_at_slot = self.model.NewBoolVar(f"{morceau}_is_at_{slot}")
                        self.model.Add(self.assignments[morceau] == self.slot_index[slot]).OnlyEnforceIf(is_at_slot)
                        self.model.Add(self.assignments[morceau] != self.slot_index[slot]).OnlyEnforceIf(is_at_slot.Not())
                        clauses.append(is_at_slot)
                    if clauses:
                        self.model.AddMaxEquality(present,clauses)
                    else:
                        self.model.Add(present == 0)
                    occupied_slots.append(present)
                    nb_slots = self.model.NewIntVar(0, len(slots), f"nb_slots_{musicien}_{jour}")
                    self.model.Add(nb_slots == sum(occupied_slots))
                    # Pénalité si le musicien est surchargé
                    is_overloaded = self.model.NewBoolVar(f"{musicien}_overloaded_{jour}")
                    self.model.Add(nb_slots > self.max_load).OnlyEnforceIf(is_overloaded)
                    self.model.Add(nb_slots <= self.max_load).OnlyEnforceIf(is_overloaded.Not())
                    penalty = self.model.NewIntVar(0, self.load_penalty, f"penalty_{musicien}_{jour}")
                    self.model.Add(penalty == self.load_penalty).OnlyEnforceIf(is_overloaded)
                    self.model.Add(penalty == 0).OnlyEnforceIf(is_overloaded.Not())
                    self.penalties.append(penalty)

    # 4th: Pénalités pour les répétitions groupées
    def add_penalites_repetitions_groupees(self):
        for musicien in self.musiciens:
            for jour, slots in self.creneaux_par_jour.items():
                morceaux_joues = [m for m in self.morceaux if musicien in self.repartition[m]]
                
                presence = []
                for slot in slots:
                    slot_idx = self.slot_index[slot]
                    var = self.model.NewBoolVar(f"{musicien}_{slot}_present")
                    
                    clauses = []
                    for morceau in morceaux_joues:
                        is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_at")
                        self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                        self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())
                        clauses.append(is_here)
                    
                    if clauses:
                        self.model.AddMaxEquality(var, clauses)
                    else:
                        self.model.Add(var == 0)
                    
                    presence.append(var)
                for i in range(len(presence) - 1):
                    bloc = self.model.NewBoolVar(f"{musicien}_{jour}_bloc_{i}")
                    self.model.AddBoolAnd([presence[i], presence[i + 1]]).OnlyEnforceIf(bloc)
                    self.model.AddBoolOr([presence[i].Not(), presence[i + 1].Not()]).OnlyEnforceIf(bloc.Not())

                    reward = self.model.NewIntVar(-self.group_bonus, 0, f"reward_bloc_{musicien}_{jour}_{i}")
                    self.model.Add(reward == -self.group_bonus).OnlyEnforceIf(bloc)
                    self.model.Add(reward == 0).OnlyEnforceIf(bloc.Not())
                    self.penalties.append(reward)  

    def define_objective(self):
        penalty_not_assigned_weight = 1000
        # 1) penalty pour non‐assignés
        for morceau in self.morceaux:
            p = self.model.NewIntVar(0, penalty_not_assigned_weight,
                                     f"penalite_non_assigne_{morceau}")
            self.model.Add(p == penalty_not_assigned_weight).OnlyEnforceIf(self.is_assigned[morceau].Not())
            self.model.Add(p == 0).OnlyEnforceIf(self.is_assigned[morceau])
            self.penalties.append(p)

        # 2) si mode "auto", on minimise ensuite T
        objective = sum(self.penalties)
        if self.mode_absence == "auto":
            W2 = 5000
            objective = objective + self.T * W2

        # 3) on inclut toutes les pénalités "non"/"peut-être"
        self.model.Minimize(objective)

    def build_model(self):
        # 1) (re)création du modèle
        self.model = cp_model.CpModel()
        # 2) définir les variables
        self.define_variables()

        # 3) ajouter toutes les contraintes
        self.add_disponibility_constraints()
        self.add_slot_constraints()
        self.add_daily_load_constraints()
        self.add_penalites_repetitions_groupees()

        # 4) définir l'objectif
        self.define_objective()


    def solve(self):
        """
        Construit le modèle (via build_model), résout, et extrait :
        - status
        - solution dict morceau->slot
        - num_unassigned
        - total_penalty
        """

        self.solver = cp_model.CpSolver()
        self.solver.parameters.max_time_in_seconds = self.generation_time_limit
        self.solver.parameters.num_search_workers = 8 # specifier le nbr de threads pour chercher la solution

        self.penalties = []
        self.musiciens_absents_force.clear()
        self.build_model()

        # --- Solve ---
        start = time.time()
        status = self.solver.Solve(self.model)
        duration = time.time() - start
        print(f"Solve status = {self.solver.StatusName(status)} en {duration:.1f}s")
        self.status = status  # stocker le statut pour l'export

        # --- Post‐processing ---
        num_unassigned = 0
        total_penalty = 0

        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            # récupérer la pénalité
            total_penalty = sum(self.solver.Value(p) for p in self.penalties)

            for morceau in self.morceaux:
                if self.solver.Value(self.is_assigned[morceau]):
                    idx = self.solver.Value(self.assignments[morceau])
                    slot = self.creneaux[idx]
                    self.solution[morceau] = slot

                    # repérer les absents forcés
                    for (mus, flag) in self.absent_participants[morceau].get(slot, []):
                        if self.solver.Value(flag):
                            self.musiciens_absents_force[morceau].add(mus)
                else:
                    num_unassigned += 1

            print(f"✅ {len(self.solution)} assignés, {num_unassigned} non-assignés, pénalité totale = {total_penalty}")
        else:
            print("⚠️ Aucune solution trouvée")

        return status, self.solution, num_unassigned, total_penalty


    def generer_planning(self):
        self.load_data()
        self.build_model()
        self.solve()

    def export_planning(self, directory=".", base_filename="planning"):
        import os
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        os.makedirs(directory, exist_ok=True)
        filename = f"{base_filename}_maybe{self.maybe_penalty}_load{self.max_load}_abs{self.seuil_absence}_timeout{self.generation_time_limit}.xlsx"
        path = os.path.join(directory, filename)
        # Constantes pour le tri et formatage (comme dans get_json_data)
        DAY_ORDER = {"Lundi":1, "Mardi":2, "Mercredi":3, "Jeudi":4,"Vendredi":5,"Samedi":6,"Dimanche":7, 
                    "LUN":1, "MAR":2, "MER":3, "JEU":4, "VEN":5, "SAM":6, "DIM":7}
        DAY_NAMES = {"LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi", "JEU":"Jeudi", 
                    "VEN":"Vendredi", "SAM":"Samedi", "DIM":"Dimanche"}

        def slot_sort_key(slot):
            jour, heures = format_slot(slot)
            reel_jour = jour.split(" ")[0]  # "Lundi 05" → "Lundi"
            return (DAY_ORDER[reel_jour], heures)

        def format_slot(slot: str) -> tuple[str,str]:
            """
            slot est au format "JOUR_DATE_START-END", ex. "LUN_05_14:00-16:00"
            Retourne (affichage_jour, affichage_heures),
            par exemple ("Lundi 05", "14:00-16:00").
            """
            try:
                jour_code, date_num, periode = slot.split("_")
            except ValueError:
                return slot, ""
            jour_aff = DAY_NAMES.get(jour_code, jour_code)
            return f"{jour_aff} {date_num}", periode

        # --- 1) DataFrame "Planning" ---
        planning_rows = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                planning_rows.append({
                    "Morceau": morceau,
                    "Jour": "Non assigné",
                    "Heures": "—",
                    "Participants": ", ".join(self.repartition.get(morceau, [])),
                    })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                planning_rows.append({
                    "Morceau": morceau,
                    "Jour": jour,
                    "Heures": heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
        df_planning = pd.DataFrame(planning_rows)
        
        # Tri du planning par jour
        if not df_planning.empty:
            df_planning["jour_order"] = df_planning["Jour"].apply(
                lambda x: DAY_ORDER.get(x.split(" ")[0], 99) if x != "Non assigné" else 100
            )
            df_planning.sort_values(["jour_order", "Heures"], inplace=True)
            df_planning.drop("jour_order", axis=1, inplace=True)

        # --- 2) DataFrames par semaine ---
        musiciens = sorted(self.musiciens)
        dispo_dfs = {}
        repart_dfs = {}

        # Pour chaque semaine détectée
        for w in self.weeks:
            # Sélection et tri des slots de la semaine w
            week_slots = [s for s in self.creneaux
                        if self.date2week[int(s.split("_")[1])] == w]
            week_slots = sorted(week_slots, key=slot_sort_key)

            # --- DataFrame Disponibilités pour cette semaine ---
            dispo_rows = []
            for slot in week_slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    row[m] = self.disponibilites.get(m, {}).get(slot, "non")
                dispo_rows.append(row)
            
            dispo_dfs[f"Dispo_Semaine_{w}"] = pd.DataFrame(dispo_rows)

            # --- DataFrame Répartition pour cette semaine ---
            repart_rows = []
            for slot in week_slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s==slot), None)
                row = {"Jour": jour, "Heures": heures, "Morceau": piece or ""}
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = self.disponibilites.get(m,{}).get(slot,"non")
                        row[m] = "Répète" if dispo=="oui" else "Répète"
                    else:
                        row[m] = ""
                repart_rows.append(row)
            
            repart_dfs[f"Repart_Semaine_{w}"] = pd.DataFrame(repart_rows)

        # --- 3) Écriture initiale via pandas ---
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_planning.to_excel(writer, sheet_name="Planning", index=False)
            
            # Écriture des onglets par semaine
            for sheet_name, df in dispo_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            for sheet_name, df in repart_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # --- 4) Re-ouverture et stylage avec openpyxl ---
        wb = load_workbook(filename)
        
        # Styles pour les disponibilités
        fill_yes    = PatternFill(fill_type="solid", fgColor="C6EFCE")  # Vert
        fill_maybe  = PatternFill(fill_type="solid", fgColor="FFEB9C")  # Jaune
        fill_no     = PatternFill(fill_type="solid", fgColor="F2DCDB")  # Rouge
        
        # Styles pour les répartitions
        fill_repete = PatternFill(fill_type="solid", fgColor="C6EFCE")  # Vert
        fill_absent = PatternFill(fill_type="solid", fgColor="FFB6C1")  # Rose
        fill_vide   = PatternFill(fill_type="solid", fgColor="D3D3D3")  # Gris
        
        # Stylage des onglets de disponibilités
        for sheet_name in dispo_dfs.keys():
            ws = wb[sheet_name]
            # Colonnes des musiciens commencent à la 3ème colonne (C)
            for row in ws.iter_rows(min_row=2, min_col=3):
                for cell in row:
                    v = (cell.value or "").strip().lower()
                    if v == "oui":
                        cell.fill = fill_yes
                    elif v in ("peut-être", "peut‐être"):
                        cell.fill = fill_maybe
                    else:
                        cell.fill = fill_no
        
        # Stylage des onglets de répartition
        for sheet_name in repart_dfs.keys():
            ws = wb[sheet_name]
            # Colonnes des musiciens commencent à la 4ème colonne (D)
            for row in ws.iter_rows(min_row=2, min_col=4):
                for cell in row:
                    v = (cell.value or "").strip()
                    if v == "Répète":
                        cell.fill = fill_repete
                    elif v == "Absent":
                        cell.fill = fill_absent
                    else:
                        cell.fill = fill_vide


        # --- Feuille Paramètres ---
        ws_params = wb.create_sheet(title="Paramètres")
        params_data = [
            ["Paramètre", "Valeur", f"Type de solution: {self.status}"],
            ["Fichier répartitions", self.repartitions_file],
            ["Fichier disponibilités", self.disponibilites_file],
            ["Pénalité maybe", self.maybe_penalty],
            ["Charge max", self.max_load],
            ["Pénalité charge", self.load_penalty],
            ["Bonus groupe", self.group_bonus],
            ["Mode absence", self.mode_absence],
            ["Seuil absence", self.seuil_absence],
            ["Temps limite génération", self.generation_time_limit]

        ]

        for row in params_data:
            ws_params.append(row)

        wb.save(path)
        return path

    def get_json_data(self):
        # pour trier et formater
        DAY_ORDER = {"Lundi":1, "Mardi":2, "Mercredi":3, "Jeudi":4,"Vendredi":5,"Samedi":6,"Dimanche":7, "LUN":1, "MAR":2, "MER":3, "JEU":4, "VEN":5, "SAM":6, "DIM":7}
        DAY_NAMES = {"LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi", "JEU":"Jeudi", "VEN":"Vendredi", "SAM":"Samedi", "DIM":"Dimanche"}

        def slot_sort_key(slot):
            jour, heures = format_slot(slot)
            reel_jour = jour.split(" ")[0]  # "Lundi 05" → "Lundi"
            return (DAY_ORDER[reel_jour], heures)

        def format_slot(slot: str) -> tuple[str,str]:
            """
            slot est au format "JOUR_DATE_START-END", ex. "LUN_05_14:00-16:00"
            Retourne (affichage_jour, affichage_heures),
            par exemple ("Lundi 05", "14:00-16:00").
            """
            # 1) découpage
            try:
                jour_code, date_num, periode = slot.split("_")
            except ValueError:
                # fallback si un slot mal formé traine
                return slot, ""
            # 2) optionnel : mappez "LUN" → "Lundi"
            jour_aff = DAY_NAMES.get(jour_code, jour_code)
            # 3) compose la chaîne complète pour l’entête "Jour"
            #    ici on peut afficher jour_aff + " " + date_num
            return f"{jour_aff} {date_num}", periode

        # 1) planning final (inchangé)
        planning = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                planning.append({
                    "Morceau": morceau,
                    "Jour":     "Non assigné",
                    "Heures":   "—",
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                planning.append({
                    "Morceau": morceau,
                    "Jour":     jour,
                    "Heures":   heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })

        # 2) DISPO & REPART pour **chaque** semaine détectée
        musiciens = sorted(self.musiciens)
        dispo_output   = {}
        repart_output  = {}

        # self.weeks contient déjà [1,2,3,…]
        for w in self.weeks:
            # sélection et tri des slots de la semaine w
            week_slots = [s for s in self.creneaux
                        if self.date2week[int(s.split("_")[1])] == w]
            week_slots = sorted(week_slots, key=slot_sort_key)

            dispo_rows  = []
            repart_rows = []

            # dispo
            for slot in week_slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    row[m] = self.disponibilites.get(m, {}).get(slot, "non")
                dispo_rows.append(row)

            # répartition
            for slot in week_slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s==slot), None)
                row = {"Jour": jour, "Heures": heures, "Morceau": piece or ""}
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = self.disponibilites.get(m,{}).get(slot,"non")
                        if dispo=="oui":
                            row[m] = "repete"
                        elif dispo =="non":
                            row[m] = "absent"
                        elif dispo == "peut-être":
                            row[m] = "maybe_absent"
                    else:
                        row[m] = "non"
                repart_rows.append(row)

            key = f"SEMAINE_{w}"
            dispo_output[key]  = dispo_rows
            repart_output[key] = repart_rows

        return {
            "planning":       planning,
            "disponibilites": dispo_output,
            "repartition":     repart_output
        }
    
