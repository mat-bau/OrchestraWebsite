"""
Description : Planificateur de répétitions pour l'Orchestraaaaa
Licence : On devrait peut-être mettre une licence hein
Anno : 43
Auteur : Mateo Bauvir
Modifié : Ajout de jours spéciaux avec tolérance d'absences
"""
import pandas as pd
import random
import time
from collections import defaultdict, Counter
from typing import Dict, List, Set, Tuple, Optional
import re

class OptimizedRepetitionScheduler:
    def __init__(self,
                 repartitions_file: str,
                 disponibilites_file: str,
                 maybe_penalty: int = 10,
                 max_load: int = 3,
                 load_penalty: int = 50,
                 group_bonus: int = 20,
                 mode_absence: str = "flexible",
                 seuil_absence: int = 2,
                 generation_time_limit: int = 30,
                 creneaux_speciaux: Optional[List[str]] = None,
                 seuil_absence_creneau_special: int = 5):
        """
        Paramètres additionnels:
        - creneaux_speciaux: Liste des créneaux où on tolère plus d'absences 
          Format: ["LUN_04_16:00-18:00", "MER_12_14:00-16:00"]
        - seuil_absence_creneau_special: Nombre d'absences tolérées pour ces créneaux spéciaux
        """
        
        self.repartitions_file = repartitions_file
        self.disponibilites_file = disponibilites_file
        self.maybe_penalty = maybe_penalty
        self.max_load = max_load
        self.load_penalty = load_penalty
        self.group_bonus = group_bonus
        self.mode_absence = mode_absence
        self.seuil_absence = seuil_absence
        self.generation_time_limit = generation_time_limit
        
        # NOUVEAUX PARAMÈTRES
        self.creneaux_speciaux = self._normaliser_creneaux_speciaux(creneaux_speciaux or [])
        self.seuil_absence_creneau_special = seuil_absence_creneau_special
        
        self.musiciens: Set[str] = set()
        self.morceaux: List[str] = []
        self.creneaux: List[str] = []
        self.weeks = []
        self.repartition: Dict[str, Set[str]] = {}
        self.repartitions_df = None
        self.disponibilites_df = None
        self.disponibilites: Dict[str, Dict[str, str]] = {}
        self.creneaux_par_jour: Dict[str, List[str]] = defaultdict(list)
        self.slot_index: Dict[str, int] = {}
        self.musiciens_absents_force: Dict[str, Set[str]] = defaultdict(set)
        self.absent_participants: Dict[str, Dict[str, List]] = defaultdict(lambda: defaultdict(list))
        
        self.assignment: Dict[str, Optional[str]] = {}
        self.conflicts: Dict[str, int] = {}
        self.solution: Dict[str, str] = {}
        self.status = None  
        self.assigned = 0
        self.notassigned = []
        
        self._conflict_cache: Dict[Tuple[str, str], int] = {}
        self._musicien_morceaux: Dict[str, List[str]] = defaultdict(list)
        
        self.max_iterations = 10000
        self.max_restarts = generation_time_limit
    
    def _normaliser_creneaux_speciaux(self, creneaux: List[str]) -> Set[str]:
        """
        Normalise les créneaux spéciaux en format standard
        Accepte plusieurs formats:
        - "LUN_04_16:00-18:00" (format complet)
        - "LUN_04_16_18" (format simplifié)
        - "Lundi 04 16:00-18:00"
        """
        DAY_CODES = {
            'lundi': 'LUN', 'lun': 'LUN',
            'mardi': 'MAR', 'mar': 'MAR',
            'mercredi': 'MER', 'mer': 'MER',
            'jeudi': 'JEU', 'jeu': 'JEU',
            'vendredi': 'VEN', 'ven': 'VEN',
            'samedi': 'SAM', 'sam': 'SAM',
            'dimanche': 'DIM', 'dim': 'DIM'
        }
        
        normalises = set()
        for creneau in creneaux:
            creneau_clean = creneau.strip().lower()
            
            # Format standard "LUN_04_16:00-18:00"
            if '_' in creneau_clean:
                parts = creneau_clean.split('_')
                if len(parts) >= 3:
                    jour_code = parts[0].upper()
                    date = f"{int(parts[1]):02d}"
                    
                    # Reconstruction de l'horaire
                    if ':' in parts[2]:
                        # Format "16:00-18:00"
                        horaire = parts[2]
                    else:
                        # Format simplifié "16_18" -> "16:00-18:00"
                        if len(parts) == 4:
                            horaire = f"{int(parts[2]):02d}:00-{int(parts[3]):02d}:00"
                        else:
                            continue
                    
                    normalises.add(f"{jour_code}_{date}_{horaire}")
            else:
                # Format "Lundi 04 16:00-18:00"
                import re
                match = re.match(r'(\w+)\s+(\d+)\s+([\d:]+\-[\d:]+)', creneau_clean)
                if match:
                    jour_nom, date, horaire = match.groups()
                    if jour_nom in DAY_CODES:
                        jour_code = DAY_CODES[jour_nom]
                        normalises.add(f"{jour_code}_{int(date):02d}_{horaire}")
        
        return normalises
    
    def _est_creneau_special(self, creneau: str) -> bool:
        """
        Vérifie si un créneau est spécial
        creneau format: "LUN_05_14:00-16:00"
        """
        return creneau in self.creneaux_speciaux
        
    def transformer_simple(self, texte: str) -> Optional[Dict]:
        """Transforme un texte de créneau en dictionnaire structuré."""
        t = texte.strip().replace("\n", " ").replace("\r", " ")
        m = re.search(r"(\w+\.)\s+(\d+).*?(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})", t)
        if not m:
            return None
        
        jour_txt, jour_chiffre, H1, M1, H2, M2 = m.groups()
        jours = {
            'lun.': 'LUN', 'mar.': 'MAR', 'mer.': 'MER',
            'jeu.': 'JEU', 'ven.': 'VEN', 'sam.': 'SAM', 'dim.': 'DIM'
        }
        jour = jours.get(jour_txt.lower(), jour_txt.upper().rstrip('.'))
        
        return {
            'jour': jour,
            'date': int(jour_chiffre),
            'h1': int(H1), 'm1': int(M1),
            'h2': int(H2), 'm2': int(M2)
        }
    
    def load_data(self):
        """Charge les données depuis les fichiers Excel."""
        
        if hasattr(self, 'repartitions_file') and self.repartitions_file:
            self.repartitions_df = pd.read_excel(self.repartitions_file)
            instrument_cols = self.repartitions_df.columns[6:]
                    
            for _, row in self.repartitions_df.iterrows():
                morceau = row['Titre']
                if pd.isna(morceau):
                    continue
                                
                has_musicians = any(not pd.isna(row[c]) for c in instrument_cols)
                if not has_musicians:
                    continue
                                
                self.morceaux.append(morceau)
                self.repartition[morceau] = set()
                            
                for col in instrument_cols:
                    cellule = row[col]
                    if pd.isna(cellule):
                        continue
                                        
                    for nom in str(cellule).split(','):
                        nom = nom.strip()
                        if nom:
                            self.musiciens.add(nom)
                            self.repartition[morceau].add(nom)
                            self._musicien_morceaux[nom].append(morceau)
        
        self.disponibilites_df = pd.read_excel(self.disponibilites_file)
        all_dates = set()
        
        if 'Nom' in self.disponibilites_df.columns:
            dispo_cols = self.disponibilites_df.columns[2:]
            
            for _, row in self.disponibilites_df.iterrows():
                musicien = str(row['Nom']).strip().title()
                if pd.isna(musicien) or not musicien:
                    continue
                                
                self.disponibilites[musicien] = {}
                            
                for col in dispo_cols:
                    info = self.transformer_simple(str(col))
                    if not info:
                        continue
                                        
                    d = info['date']
                    j = info['jour']
                    h1, m1, h2, m2 = info['h1'], info['m1'], info['h2'], info['m2']
                    all_dates.add(d)
                    start = f"{h1:02d}:{m1:02d}"
                    end = f"{h2:02d}:{m2:02d}"
                    slot = f"{j}_{d:02d}_{start}-{end}"
                                    
                    val = str(row[col]).strip().lower() if not pd.isna(row[col]) else "no"
                    self.disponibilites[musicien][slot] = val
        else:
            for idx, row in self.disponibilites_df.iterrows():
                for col_idx, col in enumerate(self.disponibilites_df.columns[2:], start=2):
                    cell_value = str(row[col]) if not pd.isna(row[col]) else ""
                    if cell_value and cell_value != "nan":
                        if any(day in cell_value for day in ['lun.', 'mar.', 'mer.', 'jeu.', 'ven.', 'sam.', 'dim.']):
                            clean_text = cell_value.replace('\n', ' ').replace('[93%]', '').strip()
                            info = self.transformer_simple(clean_text)
                            if info:
                                d = info['date']
                                j = info['jour']
                                h1, m1, h2, m2 = info['h1'], info['m1'], info['h2'], info['m2']
                                all_dates.add(d)
                                start = f"{h1:02d}:{m1:02d}"
                                end = f"{h2:02d}:{m2:02d}"
                                slot = f"{j}_{d:02d}_{start}-{end}"
                                self.creneaux.append(slot)
            
            for idx, row in self.disponibilites_df.iterrows():
                if idx <= 3:
                    continue
                    
                first_col = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
                
                if len(row) > 1 and not pd.isna(row.iloc[1]) and '@' in str(row.iloc[1]):
                    musicien = first_col.title()
                    if musicien and musicien != 'Nan':
                        self.musiciens.add(musicien)
                        self.disponibilites[musicien] = {}
                        
                        for col_idx in range(2, len(row)):
                            if col_idx - 2 < len(self.creneaux):
                                slot = self.creneaux[col_idx - 2]
                                val = str(row.iloc[col_idx]).strip().lower() if not pd.isna(row.iloc[col_idx]) else "no"
                                
                                if val in ['yes', 'oui']:
                                    self.disponibilites[musicien][slot] = "yes"
                                elif val in ['maybe', 'peut-être']:
                                    self.disponibilites[musicien][slot] = "maybe"  
                                else:
                                    self.disponibilites[musicien][slot] = "no"
            
            if not self.morceaux and self.musiciens:
                morceau_default = "Session_Planning"
                self.morceaux.append(morceau_default)
                self.repartition[morceau_default] = self.musiciens.copy()
                
                for musicien in self.musiciens:
                    self._musicien_morceaux[musicien].append(morceau_default)
        
        if self.disponibilites:
            premier = next(iter(self.disponibilites.values()))
                            
            def _key(s):
                _, dd, plage = s.split("_")
                start, _ = plage.split("-")
                h, m = start.split(":")
                return (int(dd), int(h), int(m))
            
            if 'Nom' in self.disponibilites_df.columns:
                self.creneaux = sorted(premier.keys(), key=_key)
            else:
                self.creneaux = sorted(self.creneaux, key=_key)
                
            self.slot_index = {slot: i for i, slot in enumerate(self.creneaux)}
                            
            for slot in self.creneaux:
                jour = slot.split("_")[0]
                self.creneaux_par_jour[jour].append(slot)
                    
            dates_sorted = sorted(all_dates)
            if dates_sorted:
                base = dates_sorted[0]
                self.date2week = {d: ((d - base) // 7) + 1 for d in dates_sorted}
                self.weeks = sorted(set(self.date2week.values()))
            else:
                self.weeks = []
    
    def build_model(self):
        """Construit le modèle CSP"""
        self.assignment = {}
        self.conflicts = {}
        self._conflict_cache.clear()
        self.musiciens_absents_force.clear()
        self.absent_participants.clear()
        for morceau in self.morceaux:
            self.assignment[morceau] = None
            self.conflicts[morceau] = 0
    
    def calculate_conflicts(self, morceau: str, creneau: str) -> int:
        """Calcule le nombre de conflits pour assigner un morceau à un créneau."""
        cache_key = (morceau, creneau)
        if cache_key in self._conflict_cache:
            return self._conflict_cache[cache_key]
        
        conflicts = 0
        musiciens_morceau = self.repartition[morceau]
        
        # NOUVEAU: Déterminer le seuil d'absence applicable
        est_special = self._est_creneau_special(creneau)
        seuil_actif = self.seuil_absence_creneau_special if est_special else self.seuil_absence
        
        # 1. Conflits de disponibilité
        absents = 0
        maybe_count = 0
        
        for musicien in musiciens_morceau:
            dispo = self.disponibilites.get(musicien, {}).get(creneau, "no")
            if dispo in ["non", "no"]:
                absents += 1
                if self.mode_absence == "strict" and not est_special:
                    conflicts += 10000
                else:
                    conflicts += 100
            elif dispo in ["peut-être", "maybe"]:
                maybe_count += 1
                conflicts += self.maybe_penalty
        
        # MODIFIÉ: Utiliser le seuil adapté
        if self.mode_absence != "strict" and absents > seuil_actif:
            conflicts += (absents - seuil_actif) * 10000
        
        # 2. Conflit de créneau
        for autre_morceau, autre_creneau in self.assignment.items():
            if autre_morceau != morceau and autre_creneau == creneau:
                conflicts += 100000000
        
        # 3. Conflits de charge quotidienne
        jour = creneau.split("_")[0]
        for musicien in musiciens_morceau:
            charge_jour = self._get_daily_load(musicien, jour, exclude_morceau=morceau)
            
            if any(self.assignment.get(m) == creneau for m in self._musicien_morceaux[musicien] if m != morceau):
                charge_jour += 1
            
            if charge_jour >= self.max_load:
                conflicts += self.load_penalty * (charge_jour - self.max_load + 1)
        
        # 4. Bonus pour groupements
        bonus = self._calculate_grouping_bonus(morceau, creneau)
        conflicts = max(0, conflicts - bonus)
        
        self._conflict_cache[cache_key] = conflicts
        return conflicts
    
    def _get_daily_load(self, musicien: str, jour: str, exclude_morceau: str = None) -> int:
        """Calcule la charge quotidienne d'un musicien."""
        charge = 0
        for morceau in self._musicien_morceaux[musicien]:
            if morceau == exclude_morceau:
                continue
            creneau = self.assignment.get(morceau)
            if creneau and creneau.startswith(jour + "_"):
                charge += 1
        return charge
    
    def _calculate_grouping_bonus(self, morceau: str, creneau: str) -> int:
        """Calcule le bonus de groupement pour un morceau/créneau."""
        bonus = 0
        jour = creneau.split("_")[0]
        slots_jour = self.creneaux_par_jour[jour]
        
        if creneau not in slots_jour:
            return 0
            
        creneau_idx = slots_jour.index(creneau)
        
        for musicien in self.repartition[morceau]:
            for offset in [-1, 1]:
                adj_idx = creneau_idx + offset
                if 0 <= adj_idx < len(slots_jour):
                    adj_slot = slots_jour[adj_idx]
                    
                    for autre_morceau in self._musicien_morceaux[musicien]:
                        if autre_morceau != morceau and self.assignment.get(autre_morceau) == adj_slot:
                            bonus += self.group_bonus
        
        return bonus
    
    def initialize_assignment(self):
        """Initialise l'assignation avec une solution"""
        print("Initialisation de l'assignation...")
        
        morceaux_tries = sorted(self.morceaux, 
                               key=lambda m: len(self.repartition[m]), 
                               reverse=True)
        
        for morceau in morceaux_tries:
            best_creneau = None
            min_conflicts = float('inf')
            
            for creneau in self.creneaux:
                conflicts = self.calculate_conflicts(morceau, creneau)
                if conflicts < min_conflicts:
                    min_conflicts = conflicts
                    best_creneau = creneau
            
            if best_creneau and min_conflicts < 1000:
                self.assignment[morceau] = best_creneau
            
        self._update_conflicts()
    
    def _update_conflicts(self):
        """Met à jour le compteur de conflits pour tous les morceaux."""
        self._conflict_cache.clear()
        self.conflicts = {}
        
        for morceau in self.morceaux:
            creneau = self.assignment.get(morceau)
            if creneau:
                self.conflicts[morceau] = self.calculate_conflicts(morceau, creneau)
            else:
                self.conflicts[morceau] = 10000
    
    def min_conflicts_step(self) -> bool:
        """Effectue une étape de l'algorithme min-conflicts."""
        morceaux_conflits = [(m, c) for m, c in self.conflicts.items() if c > 0]
        if not morceaux_conflits:
            return True
        
        morceaux_conflits.sort(key=lambda x: x[1], reverse=True)
        
        top_conflicted = morceaux_conflits[:min(3, len(morceaux_conflits))]
        morceau = random.choice([m for m, _ in top_conflicted])
        
        best_creneau = None
        min_conflicts = float('inf')
        
        options = self.creneaux + [None]
        
        for creneau in options:
            old_creneau = self.assignment[morceau]
            self.assignment[morceau] = creneau
            if creneau:
                conflicts = self.calculate_conflicts(morceau, creneau)
            else:
                conflicts = 500
            
            if conflicts < min_conflicts:
                min_conflicts = conflicts
                best_creneau = creneau
            
            self.assignment[morceau] = old_creneau
        
        if best_creneau != self.assignment[morceau]:
            self.assignment[morceau] = best_creneau
            self._update_conflicts()
        
        return False
    
    def solve(self):
        start_time = time.time()
        
        best_solution = None
        best_cost = float('inf')
        
        for restart in range(self.max_restarts):
            print(f"Restart {restart + 1}/{self.max_restarts}")
            
            if time.time() - start_time > self.generation_time_limit:
                print("Limite de temps atteinte")
                break
        
            self.initialize_assignment()
            
            for iteration in range(self.max_iterations):
                if time.time() - start_time > self.generation_time_limit:
                    break
                
                if self.min_conflicts_step():
                    print(f"Solution parfaite trouvée en {iteration} itérations!")
                    self.status = "OPTIMAL"
                    self._finalize_solution()
                    return
                            
            current_cost = self._calculate_total_cost()
            if current_cost < best_cost:
                best_cost = current_cost
                best_solution = dict(self.assignment)
        
        if best_solution:
            self.assignment = best_solution
            self.status = "FEASIBLE"
            print(f"Meilleure solution trouvée avec un coût de {best_cost}")
        else:
            self.status = "INFEASIBLE"
            print("Aucune solution trouvée")
        
        self._finalize_solution()
        
        duration = time.time() - start_time
        self.assigned = sum(1 for v in self.assignment.values() if v is not None)
        total_conflicts = sum(self.conflicts.values())
        
        print(f"✅ Résolution terminée en {duration:.1f}s")
        print(f"✅ {self.assigned} morceaux assignés sur {len(self.morceaux)}")
        print(f"✅ Conflits totaux: {total_conflicts}")
        
        # NOUVEAU: Afficher les créneaux spéciaux utilisés
        if self.creneaux_speciaux:
            creneaux_utilises = [c for c in self.assignment.values() 
                            if c and self._est_creneau_special(c)]
            if creneaux_utilises:
                print(f"🌟 Créneaux spéciaux : {self.creneaux_speciaux}, seuil={self.seuil_absence_creneau_special}")
    
    def _calculate_total_cost(self) -> int:
        """Calcule le coût total de la solution actuelle."""
        total = 0
        for morceau, creneau in self.assignment.items():
            if creneau:
                total += self.calculate_conflicts(morceau, creneau)
            else:
                total += 1000
        return total
    
    def _finalize_solution(self):
        """Finalise la solution et met à jour les structures compatibles."""
        self.solution = {}
        self.musiciens_absents_force.clear()
        
        for morceau, creneau in self.assignment.items():
            if creneau:
                self.solution[morceau] = creneau
                
                for musicien in self.repartition[morceau]:
                    dispo = self.disponibilites.get(musicien, {}).get(creneau, "no")
                    if dispo in ["non", "no"]:
                        self.musiciens_absents_force[morceau].add(musicien)
    
    def generer_planning(self):
        """Interface compatible avec l'ancien code."""
        self.load_data()
        self.build_model()
        self.solve()

    def export_planning(self, directory=".", base_filename="planning"):
        import os
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        os.makedirs(directory, exist_ok=True)
        filename = f"{base_filename}_maybe{self.maybe_penalty}_load{self.max_load}_abs{self.seuil_absence}_timeout{self.generation_time_limit}.xlsx"
        path = os.path.join(directory, filename)
        
        DAY_ORDER = {"Lundi":1, "Mardi":2, "Mercredi":3, "Jeudi":4,"Vendredi":5,"Samedi":6,"Dimanche":7, 
                    "LUN":1, "MAR":2, "MER":3, "JEU":4, "VEN":5, "SAM":6, "DIM":7}
        DAY_NAMES = {"LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi", "JEU":"Jeudi", 
                    "VEN":"Vendredi", "SAM":"Samedi", "DIM":"Dimanche"}

        def slot_sort_key(slot):
            jour, heures = format_slot(slot)
            reel_jour = jour.split(" ")[0]
            return (DAY_ORDER[reel_jour], heures)

        def format_slot(slot: str) -> tuple[str,str]:
            try:
                jour_code, date_num, periode = slot.split("_")
            except ValueError:
                return slot, ""
            jour_aff = DAY_NAMES.get(jour_code, jour_code)
            return f"{jour_aff} {date_num}", periode

        planning_rows = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                self.notassigned.append(morceau)
                planning_rows.append({
                    "Morceau": morceau,
                    "Jour": "Non assigné",
                    "Heures": "—",
                    "Participants": ", ".join(self.repartition.get(morceau, [])),
                    })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                planning_rows.append({
                    "Morceau": morceau,
                    "Jour": jour,
                    "Heures": heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
        df_planning = pd.DataFrame(planning_rows)
        
        if not df_planning.empty:
            df_planning["jour_order"] = df_planning["Jour"].apply(
                lambda x: DAY_ORDER.get(x.split(" ")[0], 99) if x != "Non assigné" else 100
            )
            df_planning.sort_values(["jour_order", "Heures"], inplace=True)
            df_planning.drop("jour_order", axis=1, inplace=True)

        musiciens = sorted(self.musiciens)
        dispo_dfs = {}
        repart_dfs = {}

        for w in self.weeks:
            week_slots = [s for s in self.creneaux
                        if self.date2week[int(s.split("_")[1])] == w]
            week_slots = sorted(week_slots, key=slot_sort_key)

            dispo_rows = []
            for slot in week_slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    row[m] = self.disponibilites.get(m, {}).get(slot, "no")
                dispo_rows.append(row)
            
            dispo_dfs[f"Dispo_Semaine_{w}"] = pd.DataFrame(dispo_rows)

            repart_rows = []
            for slot in week_slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s==slot), None)
                row = {"Jour": jour, "Heures": heures, "Morceau": piece or ""}
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = self.disponibilites.get(m,{}).get(slot,"no")
                        if dispo.lower() in ["oui", "yes"]:
                            row[m] = "Répète"
                        elif dispo.lower() in ["non", "no"]:
                            row[m] = "Absent"
                        elif dispo.lower() in ["peut-être", "maybe"]:
                            row[m] = "Maybe"
                        else:
                            row[m] = ""

                repart_rows.append(row)
            
            repart_dfs[f"Repart_Semaine_{w}"] = pd.DataFrame(repart_rows)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_planning.to_excel(writer, sheet_name="Planning", index=False)
            
            for sheet_name, df in dispo_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            for sheet_name, df in repart_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(filename)
        
        fill_yes    = PatternFill(fill_type="solid", fgColor="C6EFCE")
        fill_maybe  = PatternFill(fill_type="solid", fgColor="FFEB9C")
        fill_no     = PatternFill(fill_type="solid", fgColor="F2DCDB")
        
        fill_repete = PatternFill(fill_type="solid", fgColor="C6EFCE")
        fill_absent = PatternFill(fill_type="solid", fgColor="FFB6C1")
        fill_vide   = PatternFill(fill_type="solid", fgColor="D3D3D3")
        
        for sheet_name in dispo_dfs.keys():
            ws = wb[sheet_name]
            row_count = 0
            for row in ws.iter_rows(min_row=2, min_col=3):
                for cell in row:
                    v = (cell.value or "").strip().lower()
                    if v in ["oui", "yes"]:
                        cell.fill = fill_yes
                    elif v in ("peut-être", "peut‐être", "maybe"):
                        cell.fill = fill_maybe
                    else:
                        cell.fill = fill_no
                row_count += 1
        
        for sheet_name in repart_dfs.keys():
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, min_col=4):
                for cell in row:
                    v = (cell.value or "").strip()
                    if v == "Répète":
                        cell.fill = fill_repete
                    elif v == "Absent":
                        cell.fill = fill_absent
                    else:
                        cell.fill = fill_vide

        ws_params = wb.create_sheet(title="Paramètres")
        params_data = [
            ["Paramètre", "Valeur", f"Type de solution: {self.status}"],
            ["Fichier répartitions", self.repartitions_file],
            ["Fichier disponibilités", self.disponibilites_file],
            ["Pénalité maybe", self.maybe_penalty],
            ["Charge max", self.max_load],
            ["Pénalité charge", self.load_penalty],
            ["Bonus groupe", self.group_bonus],
            ["Mode absence", self.mode_absence],
            ["Seuil absence", self.seuil_absence],
            ["Créneaux spéciaux", ", ".join(self.creneaux_speciaux) if self.creneaux_speciaux else "Aucun"],
            ["Seuil absence créneaux spéciaux", self.seuil_absence_creneau_special],
            ["Temps limite génération", self.generation_time_limit]
        ]

        for row in params_data:
            ws_params.append(row)

        wb.save(path)
        return path

    def get_json_data(self):
        DAY_ORDER = {"Lundi":1, "Mardi":2, "Mercredi":3, "Jeudi":4,"Vendredi":5,"Samedi":6,"Dimanche":7, "LUN":1, "MAR":2, "MER":3, "JEU":4, "VEN":5, "SAM":6, "DIM":7}
        DAY_NAMES = {"LUN":"Lundi", "MAR":"Mardi", "MER":"Mercredi", "JEU":"Jeudi", "VEN":"Vendredi", "SAM":"Samedi", "DIM":"Dimanche"}

        def slot_sort_key(slot):
            jour, heures = format_slot(slot)
            reel_jour = jour.split(" ")[0]
            return (DAY_ORDER[reel_jour], heures)

        def format_slot(slot: str) -> tuple[str,str]:
            try:
                jour_code, date_num, periode = slot.split("_")
            except ValueError:
                return slot, ""
            jour_aff = DAY_NAMES.get(jour_code, jour_code)
            return f"{jour_aff} {date_num}", periode

        planning = []
        for morceau in self.morceaux:
            if morceau not in self.solution:
                planning.append({
                    "Morceau": morceau,
                    "Jour":     "Non assigné",
                    "Heures":   "—",
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
            else:
                slot = self.solution[morceau]
                jour, heures = format_slot(slot)
                planning.append({
                    "Morceau": morceau,
                    "Jour":     jour,
                    "Heures":   heures,
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })

        musiciens = sorted(self.musiciens)
        dispo_output   = {}
        repart_output  = {}

        for w in self.weeks:
            week_slots = [s for s in self.creneaux
                        if self.date2week[int(s.split("_")[1])] == w]
            week_slots = sorted(week_slots, key=slot_sort_key)

            dispo_rows  = []
            repart_rows = []

            for slot in week_slots:
                jour, heures = format_slot(slot)
                row = {"Jour": jour, "Heures": heures}
                for m in musiciens:
                    row[m] = self.disponibilites.get(m, {}).get(slot, "no")
                dispo_rows.append(row)

            for slot in week_slots:
                jour, heures = format_slot(slot)
                piece = next((p for p,s in self.solution.items() if s==slot), None)
                row = {"Jour": jour, "Heures": heures, "Morceau": piece or ""}
                for m in musiciens:
                    if piece and m in self.repartition.get(piece, []):
                        dispo = self.disponibilites.get(m,{}).get(slot,"no")
                        if dispo.lower() in ["oui", "yes"]:
                            row[m] = "repete"
                        elif dispo.lower() in ["non", "no"]:
                            row[m] = "absent"
                        elif dispo.lower() in ["peut-être", "maybe"]:
                            row[m] = "maybe_absent"
                    else:
                        row[m] = "no"
                repart_rows.append(row)

            key = f"SEMAINE_{w}"
            dispo_output[key]  = dispo_rows
            repart_output[key] = repart_rows

        return {
            "planning":       planning,
            "disponibilites": dispo_output,
            "repartition":     repart_output,
            "assigned": self.assigned,
            "total": len(self.morceaux),
            "notassigned": self.notassigned
        }